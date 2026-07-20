"""Excel/CSV import — the counterpart to ``export.py``.

Column headers are matched case-insensitively against the collection's field
*labels* and *keys*; a column named "Art" selects/creates the ``ItemType``.
Because the Excel export writes exactly these headers, export → edit → import
round-trips. Image/file columns (URLs in the export) cannot be imported and are
reported as ignored; unparsable cells are dropped with a warning, the row is
still imported.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, time

from django.utils.translation import gettext as _

from .dynamic_forms import FILE_TYPES
from .models import FieldType, Item, ItemType

_ART_HEADERS = {'art', 'typ', 'type'}
_SILENT_HEADERS = {'', 'id'}
_TRUE_WORDS = {'ja', 'yes', 'true', 'wahr', '1', 'x'}
_FALSE_WORDS = {'nein', 'no', 'false', 'falsch', '0', ''}


def read_rows(uploaded_file) -> list[list]:
    """Return the file's rows as lists of cell values (first row = header)."""
    if uploaded_file.name.lower().endswith('.csv'):
        text = uploaded_file.read().decode('utf-8-sig', errors='replace')
        try:
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=';,\t')
        except csv.Error:
            dialect = csv.excel
        return [row for row in csv.reader(io.StringIO(text), dialect)]
    from openpyxl import load_workbook
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    return [list(row) for row in workbook.active.iter_rows(values_only=True)]


def _parse_date(raw) -> str:
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    text = str(raw).strip()
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d.%m.%y'):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(text)


def parse_value(fd, raw):
    """Convert one cell to the JSON value for ``fd``. Raises ValueError."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    t = fd.field_type

    if t in (FieldType.NUMBER, FieldType.YEAR):
        return int(float(str(raw).replace(',', '.')))
    if t in (FieldType.DECIMAL, FieldType.PRICE):
        return float(str(raw).replace(',', '.'))
    if t == FieldType.BOOLEAN:
        if isinstance(raw, bool):
            return raw
        text = str(raw).strip().lower()
        if text in _TRUE_WORDS:
            return True
        if text in _FALSE_WORDS:
            return False
        raise ValueError(text)
    if t == FieldType.DATE:
        return _parse_date(raw)
    if t == FieldType.TIME:
        if isinstance(raw, (datetime, time)):
            return raw.strftime('%H:%M')
        return str(raw).strip()
    if t in (FieldType.CHOICE, FieldType.MULTICHOICE):
        choices = {str(c).lower(): str(c) for c in (fd.config or {}).get('choices', [])}
        if t == FieldType.CHOICE:
            value = choices.get(str(raw).strip().lower())
            if value is None:
                raise ValueError(str(raw))
            return value
        parts = [p.strip() for p in re.split(r'[;,]', str(raw)) if p.strip()]
        matched = [choices[p.lower()] for p in parts if p.lower() in choices]
        if parts and not matched:
            raise ValueError(str(raw))
        return matched
    if t == FieldType.URL:
        # Mirror the form field's behaviour: accept http(s), add a scheme to
        # bare domains, refuse anything else (javascript:, data:, …).
        text = str(raw).strip()
        lowered = text.lower()
        if lowered.startswith(('http://', 'https://')):
            return text
        if '://' in text or lowered.startswith(('javascript:', 'data:', 'vbscript:')):
            raise ValueError(text)
        return 'https://' + text
    # text/textarea/isbn/barcode/email/datetime: store the trimmed string.
    return str(raw).strip()


def import_table(collection, rows: list[list], user=None) -> dict:
    """Import ``rows`` (header + data) into ``collection``.

    Returns ``{'created': int, 'ignored_columns': [..], 'warnings': [..]}``.
    """
    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    by_name = {}
    for fd in collection.fields.all():
        by_name[fd.label.strip().lower()] = fd
        by_name[fd.key.lower()] = fd

    mapping, ignored, art_index = [], [], None
    for index, name in enumerate(header):
        lowered = name.lower()
        if lowered in _ART_HEADERS:
            art_index = index
        elif lowered in _SILENT_HEADERS:
            continue
        elif lowered in by_name and by_name[lowered].field_type not in FILE_TYPES:
            mapping.append((index, by_name[lowered]))
        else:
            ignored.append(name)

    created, warnings = 0, []
    for row_number, row in enumerate(rows[1:], start=2):
        cells = list(row) + [None] * (len(header) - len(row))
        values = {}
        for index, fd in mapping:
            try:
                value = parse_value(fd, cells[index])
            except (ValueError, TypeError):
                warnings.append(_('Zeile %(row)s: Wert „%(value)s“ für „%(field)s“ übersprungen.')
                                % {'row': row_number, 'value': cells[index], 'field': fd.label})
                continue
            if value is not None:
                values[fd.key] = value
        item_type = None
        art_name = str(cells[art_index] or '').strip() if art_index is not None else ''
        if art_name:
            item_type, _created = ItemType.objects.get_or_create(
                collection=collection, name=art_name[:120])
        if not values and not item_type:
            continue  # fully empty row
        Item.objects.create(collection=collection, item_type=item_type,
                            values=values, created_by=user)
        created += 1
    return {'created': created, 'ignored_columns': ignored, 'warnings': warnings}
