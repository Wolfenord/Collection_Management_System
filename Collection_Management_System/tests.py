"""Automated regression tests for the Collection Management System.

Covers the full feature set: auth, collections + row-level permissions, dynamic
fields, item CRUD with per-"Art" required fields and file uploads, filtering,
QR/barcode, Excel export, statistics, sharing and "copy structure from template".
"""

import io
import json
import tempfile
from unittest import mock

from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase as DjangoTestCase
from django.test import override_settings
from django.urls import reverse
from PIL import Image


class TestCase(DjangoTestCase):
    """Project TestCase: start every test with an empty cache so brute-force
    and lookup rate-limit counters (accounts.throttling) never leak between
    tests (SQLite reuses primary keys after the per-test rollback)."""

    def _pre_setup(self):
        super()._pre_setup()
        cache.clear()

from . import imports, lookup_providers
from .models import (
    Collection,
    CollectionShare,
    FieldDefinition,
    FieldType,
    Item,
    ItemType,
    Loan,
)
from .services import create_default_fields

User = get_user_model()

MEDIA = tempfile.mkdtemp()


def make_png(color='red') -> bytes:
    buf = io.BytesIO()
    Image.new('RGB', (8, 8), color).save(buf, 'PNG')
    return buf.getvalue()


def add_field(collection, key, label, field_type, order=0, required=False, config=None):
    return FieldDefinition.objects.create(
        collection=collection, key=key, label=label, field_type=field_type,
        order=order, required=required, config=config or {},
    )


class AuthTests(TestCase):
    def test_register_creates_pending_user_awaiting_approval(self):
        # Registration now puts the account on the admin-approval whitelist:
        # it is created locked and the user is NOT logged in. See
        # accounts/tests.py for the full approval-workflow coverage.
        resp = self.client.post(reverse('register'), {
            'username': 'neo', 'email': 'neo@e.de',
            'password1': 'Sehr-Sicher-123', 'password2': 'Sehr-Sicher-123',
        })
        self.assertRedirects(resp, reverse('login'))
        user = User.objects.get(username='neo')
        self.assertFalse(user.is_active)
        self.assertEqual(user.approval_status, User.APPROVAL_PENDING)
        self.assertNotIn('_auth_user_id', self.client.session)

    def test_dashboard_requires_login(self):
        resp = self.client.get(reverse('dashboard'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse('login'), resp['Location'])


class CollectionTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.stranger = User.objects.create_user('stranger', 's@e.de', 'pw')

    def test_create_seeds_default_fields(self):
        self.client.force_login(self.owner)
        resp = self.client.post(reverse('collection_create'), {'name': 'Filme', 'description': ''})
        col = Collection.objects.get(name='Filme')
        self.assertRedirects(resp, reverse('collection_detail', args=[col.pk]))
        self.assertEqual(col.owner, self.owner)
        self.assertEqual(
            sorted(col.fields.values_list('key', flat=True)),
            ['beleg', 'bild', 'kaufdatum', 'name', 'ort', 'preis'],
        )

    def test_stranger_cannot_view(self):
        col = Collection.objects.create(owner=self.owner, name='Privat')
        self.client.force_login(self.stranger)
        resp = self.client.get(reverse('collection_detail', args=[col.pk]))
        self.assertEqual(resp.status_code, 403)


class FieldTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        self.client.force_login(self.owner)

    def test_create_choice_field_stores_choices(self):
        self.client.post(reverse('field_create', args=[self.col.pk]), {
            'label': 'Genre', 'key': '', 'field_type': 'choice', 'order': 1,
            'choices_text': 'Action\nDrama',
        })
        field = FieldDefinition.objects.get(collection=self.col, key='genre')
        self.assertEqual(field.config.get('choices'), ['Action', 'Drama'])

    def test_delete_field_removes_values_everywhere(self):
        field = add_field(self.col, 'genre', 'Genre', FieldType.CHOICE, config={'choices': ['A']})
        item = Item.objects.create(collection=self.col, values={'genre': 'A', 'name': 'X'})
        self.client.post(reverse('field_delete', args=[self.col.pk, field.pk]))
        item.refresh_from_db()
        self.assertNotIn('genre', item.values)
        self.assertFalse(FieldDefinition.objects.filter(pk=field.pk).exists())


class FieldReorderTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        self.a = add_field(self.col, 'a', 'A', FieldType.TEXT, order=0)
        self.b = add_field(self.col, 'b', 'B', FieldType.TEXT, order=1)
        self.c = add_field(self.col, 'c', 'C', FieldType.TEXT, order=2)
        self.url = reverse('field_reorder', args=[self.col.pk])

    def test_reorder_persists_new_order(self):
        self.client.force_login(self.owner)
        resp = self.client.post(
            self.url,
            data=json.dumps({'order': [str(self.c.pk), str(self.a.pk), str(self.b.pk)]}),
            content_type='application/json',
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(list(self.col.fields.values_list('key', flat=True)), ['c', 'a', 'b'])

    def test_get_not_allowed(self):
        self.client.force_login(self.owner)
        self.assertEqual(self.client.get(self.url).status_code, 405)

    def test_viewer_cannot_reorder(self):
        CollectionShare.objects.create(collection=self.col, user=self.viewer, permission='view')
        self.client.force_login(self.viewer)
        resp = self.client.post(self.url, data=json.dumps({'order': []}), content_type='application/json')
        self.assertEqual(resp.status_code, 403)


@override_settings(MEDIA_ROOT=MEDIA)
class ItemTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        create_default_fields(self.col)
        add_field(self.col, 'genre', 'Genre', FieldType.CHOICE, order=10, config={'choices': ['Action', 'Drama']})
        self.dvd = ItemType.objects.create(collection=self.col, name='DVD')
        self.dvd.required_fields.set([self.col.fields.get(key='genre')])
        self.client.force_login(self.owner)

    def test_per_type_required_blocks_missing_value(self):
        resp = self.client.post(reverse('item_create', args=[self.col.pk]),
                                {'__item_type': str(self.dvd.pk), 'name': 'Matrix'})
        self.assertEqual(resp.status_code, 200)  # re-render with error
        self.assertEqual(Item.objects.count(), 0)

    def test_create_with_file_upload_creates_asset(self):
        resp = self.client.post(reverse('item_create', args=[self.col.pk]), {
            '__item_type': str(self.dvd.pk), 'name': 'Matrix', 'genre': 'Action',
            'preis': '9.99', 'kaufdatum': '2024-05-01',
            'bild': SimpleUploadedFile('p.png', make_png(), content_type='image/png'),
        })
        self.assertEqual(resp.status_code, 302)
        item = Item.objects.get()
        self.assertEqual(item.values['name'], 'Matrix')
        self.assertEqual(item.values['preis'], 9.99)  # stored as number
        self.assertTrue(item.assets.filter(field_key='bild').exists())
        self.assertIn('url', item.values['bild'])

    def test_edit_keeps_existing_file_when_not_replaced(self):
        item = Item.objects.create(collection=self.col, item_type=self.dvd,
                                   values={'name': 'Matrix', 'genre': 'Action'})
        item.assets.create(field_key='bild', file=SimpleUploadedFile('p.png', make_png()))
        resp = self.client.post(reverse('item_edit', args=[self.col.pk, item.pk]),
                                {'__item_type': str(self.dvd.pk), 'name': 'Reloaded', 'genre': 'Drama'})
        self.assertEqual(resp.status_code, 302)
        item.refresh_from_db()
        self.assertEqual(item.values['name'], 'Reloaded')
        self.assertTrue(item.assets.filter(field_key='bild').exists())

    def test_viewer_cannot_create(self):
        CollectionShare.objects.create(collection=self.col, user=self.viewer, permission='view')
        self.client.force_login(self.viewer)
        resp = self.client.post(reverse('item_create', args=[self.col.pk]), {'name': 'x'})
        self.assertEqual(resp.status_code, 403)

    def test_item_form_offers_camera_capture_for_file_fields(self):
        resp = self.client.get(reverse('item_create', args=[self.col.pk]))
        self.assertContains(resp, 'data-capture="image"')  # Bild
        self.assertContains(resp, 'data-capture="file"')  # Beleg
        self.assertContains(resp, 'id="captureModal"')
        self.assertContains(resp, 'js/capture.js')

    def test_autofill_cover_url_is_downloaded_into_asset(self):
        with mock.patch.object(lookup_providers, 'fetch_cover', return_value=(b'IMG', 'jpg')) as m:
            resp = self.client.post(reverse('item_create', args=[self.col.pk]), {
                '__item_type': str(self.dvd.pk), 'name': 'Matrix', 'genre': 'Action',
                'bild__cover_url': 'https://portal.dnb.de/opac/mvb/cover?isbn=9783608963762',
            })
        self.assertEqual(resp.status_code, 302)
        m.assert_called_once()
        item = Item.objects.get()
        self.assertTrue(item.assets.filter(field_key='bild').exists())
        self.assertIn('url', item.values['bild'])

    def test_uploaded_file_wins_over_autofill_cover(self):
        with mock.patch.object(lookup_providers, 'fetch_cover') as m:
            resp = self.client.post(reverse('item_create', args=[self.col.pk]), {
                '__item_type': str(self.dvd.pk), 'name': 'Matrix', 'genre': 'Action',
                'bild__cover_url': 'https://portal.dnb.de/opac/mvb/cover?isbn=1',
                'bild': SimpleUploadedFile('eigen.png', make_png(), content_type='image/png'),
            })
        self.assertEqual(resp.status_code, 302)
        m.assert_not_called()
        self.assertEqual(Item.objects.get().assets.get(field_key='bild').original_name, 'eigen.png')


class FilterTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        add_field(self.col, 'genre', 'Genre', FieldType.CHOICE, order=1, config={'choices': ['Action', 'Drama']})
        add_field(self.col, 'jahr', 'Jahr', FieldType.YEAR, order=2)
        Item.objects.create(collection=self.col, values={'name': 'Matrix', 'genre': 'Action', 'jahr': 1999})
        Item.objects.create(collection=self.col, values={'name': 'Amelie', 'genre': 'Drama', 'jahr': 2001})
        self.client.force_login(self.owner)

    def _count(self, query):
        return self.client.get(reverse('collection_detail', args=[self.col.pk]), query).context['result_count']

    def test_text_choice_and_range_filters(self):
        self.assertEqual(self._count({'q': 'matrix'}), 1)
        self.assertEqual(self._count({'f_genre': 'Action'}), 1)
        self.assertEqual(self._count({'min_jahr': '2000'}), 1)
        self.assertEqual(self._count({'min_jahr': '1990', 'max_jahr': '2010'}), 2)

    def test_multichoice_membership_including_non_ascii(self):
        add_field(self.col, 'tags', 'Tags', FieldType.MULTICHOICE, order=3,
                  config={'choices': ['Action', 'Komödie']})
        Item.objects.create(collection=self.col, values={'name': 'X', 'tags': ['Action', 'Komödie']})
        Item.objects.create(collection=self.col, values={'name': 'Y', 'tags': ['Komödie']})
        self.assertEqual(self._count({'f_tags': 'Action'}), 1)
        self.assertEqual(self._count({'f_tags': 'Komödie'}), 2)  # non-ASCII must match too


class CodeTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.stranger = User.objects.create_user('stranger', 's@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        add_field(self.col, 'name', 'Name', FieldType.TEXT)
        self.item = Item.objects.create(collection=self.col, values={'name': 'Matrix'})

    def test_qr_and_barcode_return_png(self):
        self.client.force_login(self.owner)
        for url in [
            reverse('collection_qr', args=[self.col.pk]),
            reverse('item_qr', args=[self.col.pk, self.item.pk]),
            reverse('item_barcode', args=[self.col.pk, self.item.pk]),
        ]:
            resp = self.client.get(url)
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp['Content-Type'], 'image/png')
            self.assertTrue(resp.content.startswith(b'\x89PNG'))

    def test_codes_require_access(self):
        self.client.force_login(self.stranger)
        resp = self.client.get(reverse('item_qr', args=[self.col.pk, self.item.pk]))
        self.assertEqual(resp.status_code, 403)


class LabelTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.stranger = User.objects.create_user('stranger', 's@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        add_field(self.col, 'name', 'Name', FieldType.TEXT)
        Item.objects.create(collection=self.col, values={'name': 'Matrix'})

    def test_labels_pdf_download(self):
        self.client.force_login(self.owner)
        resp = self.client.get(reverse('collection_labels', args=[self.col.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertTrue(resp.content.startswith(b'%PDF'))
        self.assertIn('attachment', resp['Content-Disposition'])

    def test_labels_handles_empty_filter_result(self):
        self.client.force_login(self.owner)
        resp = self.client.get(reverse('collection_labels', args=[self.col.pk]), {'f_name': 'nichts'})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.content.startswith(b'%PDF'))

    def test_labels_require_access(self):
        self.client.force_login(self.stranger)
        resp = self.client.get(reverse('collection_labels', args=[self.col.pk]))
        self.assertEqual(resp.status_code, 403)


class ExportTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        add_field(self.col, 'genre', 'Genre', FieldType.CHOICE, order=1, config={'choices': ['Action', 'Drama']})
        Item.objects.create(collection=self.col, values={'name': 'Matrix', 'genre': 'Action'})
        Item.objects.create(collection=self.col, values={'name': 'Amelie', 'genre': 'Drama'})
        self.client.force_login(self.owner)

    def test_export_headers_and_filter(self):
        from openpyxl import load_workbook
        resp = self.client.get(reverse('collection_export', args=[self.col.pk]))
        self.assertEqual(resp.status_code, 200)
        ws = load_workbook(io.BytesIO(resp.content)).active
        self.assertEqual([c.value for c in ws[1]], ['ID', 'Art', 'Name', 'Genre'])
        self.assertEqual(ws.max_row, 3)  # header + 2 items

        resp = self.client.get(reverse('collection_export', args=[self.col.pk]), {'f_genre': 'Action'})
        ws = load_workbook(io.BytesIO(resp.content)).active
        self.assertEqual(ws.max_row, 2)  # header + 1 item


class StatisticsTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        add_field(self.col, 'preis', 'Preis', FieldType.PRICE, order=0, config={'currency': 'EUR'})
        add_field(self.col, 'genre', 'Genre', FieldType.CHOICE, order=1, config={'choices': ['Action', 'Drama']})
        Item.objects.create(collection=self.col, values={'preis': 10.0, 'genre': 'Action'})
        Item.objects.create(collection=self.col, values={'preis': 20.0, 'genre': 'Action'})
        Item.objects.create(collection=self.col, values={'preis': 15.0, 'genre': 'Drama'})
        self.client.force_login(self.owner)

    def test_numeric_total_and_filtered(self):
        resp = self.client.get(reverse('collection_statistics', args=[self.col.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['total_value'], 45.0)
        self.assertEqual(resp.context['stats']['total_items'], 3)
        price = next(n for n in resp.context['stats']['numeric'] if n['field'].key == 'preis')
        self.assertEqual((price['sum'], price['avg'], price['min'], price['max']), (45.0, 15.0, 10.0, 20.0))

        resp = self.client.get(reverse('collection_statistics', args=[self.col.pk]), {'f_genre': 'Action'})
        self.assertEqual(resp.context['total_value'], 30.0)


class ShareTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.anna = User.objects.create_user('anna', 'anna@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        add_field(self.col, 'name', 'Name', FieldType.TEXT)
        self.url = reverse('collection_shares', args=[self.col.pk])

    def test_share_grants_view_then_upgrade_to_edit(self):
        self.client.force_login(self.owner)
        self.client.post(self.url, {'identifier': 'anna@e.de', 'permission': 'view'})
        self.assertEqual(self.col.shares.count(), 1)

        self.client.force_login(self.anna)
        self.assertEqual(self.client.get(reverse('collection_detail', args=[self.col.pk])).status_code, 200)
        self.assertEqual(self.client.post(reverse('item_create', args=[self.col.pk]), {'name': 'x'}).status_code, 403)

        # upgrade by username -> still a single share, now editable
        self.client.force_login(self.owner)
        self.client.post(self.url, {'identifier': 'anna', 'permission': 'edit'})
        self.assertEqual(self.col.shares.count(), 1)
        self.client.force_login(self.anna)
        self.assertEqual(self.client.post(reverse('item_create', args=[self.col.pk]), {'name': 'x'}).status_code, 302)

    def test_non_owner_cannot_manage_shares(self):
        self.client.force_login(self.anna)
        self.assertEqual(self.client.get(self.url).status_code, 403)

    def test_share_errors(self):
        self.client.force_login(self.owner)
        resp = self.client.post(self.url, {'identifier': 'nobody', 'permission': 'view'})
        self.assertTrue(resp.context['form'].errors)
        resp = self.client.post(self.url, {'identifier': 'owner', 'permission': 'view'})
        self.assertTrue(resp.context['form'].errors)


class CopyStructureTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.other = User.objects.create_user('other', 'p@e.de', 'pw')
        self.src = Collection.objects.create(owner=self.owner, name='Vorlage')
        add_field(self.src, 'name', 'Name', FieldType.TEXT, order=0)
        add_field(self.src, 'genre', 'Genre', FieldType.CHOICE, order=1, config={'choices': ['A']})
        dvd = ItemType.objects.create(collection=self.src, name='DVD')
        dvd.required_fields.set([self.src.fields.get(key='genre')])
        Item.objects.create(collection=self.src, values={'name': 'Matrix'})
        self.client.force_login(self.owner)

    def test_copy_fields_and_types_not_items(self):
        self.client.post(reverse('collection_create'),
                         {'name': 'Neu', 'description': '', 'template': str(self.src.pk)})
        neu = Collection.objects.get(name='Neu')
        self.assertEqual(sorted(neu.fields.values_list('key', flat=True)), ['genre', 'name'])
        new_type = neu.item_types.get(name='DVD')
        self.assertEqual(list(new_type.required_fields.values_list('key', flat=True)), ['genre'])
        self.assertEqual(new_type.required_fields.first().collection_id, neu.pk)
        self.assertEqual(neu.items.count(), 0)

    def test_cannot_use_inaccessible_collection_as_template(self):
        private = Collection.objects.create(owner=self.other, name='Privat')
        resp = self.client.post(reverse('collection_create'),
                                {'name': 'Hack', 'template': str(private.pk)})
        self.assertTrue(resp.context['form'].errors.get('template'))
        self.assertFalse(Collection.objects.filter(name='Hack').exists())


# --- External-database auto-fill (QR/ISBN scan) -------------------------------

OPENLIBRARY_PAYLOAD = {
    'ISBN:9780132350884': {
        'title': 'Clean Code',
        'authors': [{'name': 'Robert C. Martin'}],
        'publishers': [{'name': 'Prentice Hall'}],
        'publish_date': '2008',
        'number_of_pages': 464,
        'subjects': [{'name': 'Software engineering'}],
        'identifiers': {'isbn_13': ['9780132350884']},
        'cover': {'large': 'https://covers.example/large.jpg'},
    }
}

GOOGLE_PAYLOAD = {
    'totalItems': 1,
    'items': [{
        'volumeInfo': {
            'title': 'Clean Code',
            'authors': ['Robert C. Martin'],
            'publisher': 'Prentice Hall',
            'publishedDate': '2008-08-01',
            'pageCount': 464,
            'description': 'A handbook of agile software craftsmanship.',
            'categories': ['Computers'],
            'language': 'en',
            'industryIdentifiers': [{'type': 'ISBN_13', 'identifier': '9780132350884'}],
            'imageLinks': {'thumbnail': 'https://books.example/thumb.jpg'},
        }
    }],
}


DNB_PAYLOAD = """<?xml version="1.0" encoding="UTF-8"?>
<searchRetrieveResponse xmlns="http://www.loc.gov/zing/srw/"><version>1.1</version>
<numberOfRecords>1</numberOfRecords><records><record><recordSchema>oai_dc</recordSchema>
<recordPacking>xml</recordPacking><recordData>
<dc xmlns="http://www.openarchives.org/OAI/2.0/oai_dc/" xmlns:dc="http://purl.org/dc/elements/1.1/"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>[Brief answers to the big questions] ; Kurze Antworten auf große Fragen / Stephen Hawking</dc:title>
  <dc:creator>Hawking, Stephen W. [Verfasser]</dc:creator>
  <dc:creator>Kober, Hainer [Übersetzer]</dc:creator>
  <dc:publisher>Stuttgart : Klett-Cotta</dc:publisher>
  <dc:date>2018</dc:date>
  <dc:language>ger</dc:language>
  <dc:identifier xsi:type="tel:ISBN">978-3-608-96376-2 Festeinband : circa EUR 16.00 (DE)</dc:identifier>
  <dc:identifier xsi:type="tel:ISBN">3-608-96376-6</dc:identifier>
  <dc:subject>500 Naturwissenschaften</dc:subject>
  <dc:format>252 Seiten</dc:format>
</dc></recordData></record></records></searchRetrieveResponse>"""

DNB_EMPTY_PAYLOAD = """<?xml version="1.0" encoding="UTF-8"?>
<searchRetrieveResponse xmlns="http://www.loc.gov/zing/srw/">
<version>1.1</version><numberOfRecords>0</numberOfRecords></searchRetrieveResponse>"""

DNB_SEARCH_PAYLOAD = """<?xml version="1.0" encoding="UTF-8"?>
<searchRetrieveResponse xmlns="http://www.loc.gov/zing/srw/"><version>1.1</version>
<numberOfRecords>2</numberOfRecords><records>
<record><recordData>
<dc xmlns="http://www.openarchives.org/OAI/2.0/oai_dc/" xmlns:dc="http://purl.org/dc/elements/1.1/"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>Kurze Antworten auf große Fragen / Stephen Hawking</dc:title>
  <dc:creator>Hawking, Stephen W. [Verfasser]</dc:creator>
  <dc:date>2018</dc:date>
  <dc:identifier xsi:type="tel:ISBN">978-3-608-96376-2</dc:identifier>
</dc></recordData></record>
<record><recordData>
<dc xmlns="http://www.openarchives.org/OAI/2.0/oai_dc/" xmlns:dc="http://purl.org/dc/elements/1.1/"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>Eine kurze Geschichte der Zeit / Stephen Hawking</dc:title>
  <dc:creator>Hawking, Stephen W. [Verfasser]</dc:creator>
  <dc:date>1988</dc:date>
  <dc:identifier xsi:type="tel:ISBN">3-498-02884-7</dc:identifier>
</dc></recordData></record>
</records></searchRetrieveResponse>"""

OPENLIBRARY_SEARCH_PAYLOAD = {
    'docs': [
        {'title': 'Clean Code', 'author_name': ['Robert C. Martin'], 'first_publish_year': 2008,
         'publisher': ['Prentice Hall'], 'isbn': ['0132350882', '9780132350884'],
         'cover_i': 123, 'number_of_pages_median': 464},
        {'title': 'The Clean Coder', 'author_name': ['Robert C. Martin'], 'first_publish_year': 2011},
    ],
}


class ProviderTests(TestCase):
    """The pre-configured providers normalise external data to the shared vocab."""

    def test_openlibrary_normalises_and_strips_isbn_hyphens(self):
        with mock.patch.object(lookup_providers, '_http_get_json', return_value=OPENLIBRARY_PAYLOAD) as m:
            data = lookup_providers.get_provider('openlibrary').fetch('978-0-13-235088-4')
        # hyphenated ISBN must reach the API as bare digits
        self.assertIn('9780132350884', m.call_args[0][0])
        self.assertEqual(data['title'], 'Clean Code')
        self.assertEqual(data['authors'], 'Robert C. Martin')
        self.assertEqual(data['publisher'], 'Prentice Hall')
        self.assertEqual(data['year'], '2008')
        self.assertEqual(data['pages'], 464)
        self.assertEqual(data['isbn'], '9780132350884')
        self.assertTrue(data['cover_url'].endswith('large.jpg'))

    def test_googlebooks_normalises_and_adds_description(self):
        with mock.patch.object(lookup_providers, '_http_get_json', return_value=GOOGLE_PAYLOAD):
            data = lookup_providers.get_provider('googlebooks').fetch('9780132350884')
        self.assertEqual(data['title'], 'Clean Code')
        self.assertEqual(data['year'], '2008')
        self.assertEqual(data['language'], 'en')
        self.assertIn('agile', data['description'])

    def test_dnb_parses_sru_xml_and_cleans_german_records(self):
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_PAYLOAD) as m:
            data = lookup_providers.get_provider('dnb').fetch('978-3-608-96376-2')
        self.assertIn('9783608963762', m.call_args[0][0])  # bare digits reach the API
        self.assertEqual(data['title'], 'Kurze Antworten auf große Fragen')
        self.assertEqual(data['authors'], 'Stephen W. Hawking')  # only [Verfasser], reordered
        self.assertEqual(data['publisher'], 'Klett-Cotta')  # place stripped
        self.assertEqual(data['year'], '2018')
        self.assertEqual(data['pages'], 252)
        self.assertEqual(data['language'], 'Deutsch')
        self.assertEqual(data['isbn'], '9783608963762')  # ISBN-13 preferred, price stripped
        self.assertEqual(data['categories'], 'Naturwissenschaften')  # DDC number stripped
        self.assertIn('9783608963762', data['cover_url'])

    def test_dnb_no_match_and_bad_xml_return_empty(self):
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_EMPTY_PAYLOAD):
            self.assertEqual(lookup_providers.get_provider('dnb').fetch('9783608963762'), {})
        with mock.patch.object(lookup_providers, '_http_get_text', return_value='not xml <'):
            self.assertEqual(lookup_providers.get_provider('dnb').fetch('9783608963762'), {})
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=None):
            self.assertEqual(lookup_providers.get_provider('dnb').fetch('9783608963762'), {})

    def test_auto_provider_merges_chain_results(self):
        # DNB answers first (its values win), Google adds the description DNB lacks.
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json',
                                  side_effect=lambda url: GOOGLE_PAYLOAD if 'googleapis' in url else {}):
            data = lookup_providers.get_provider('auto').fetch('9783608963762')
        self.assertEqual(data['title'], 'Kurze Antworten auf große Fragen')
        self.assertIn('agile', data['description'])

    def test_auto_provider_falls_back_when_first_source_misses(self):
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_EMPTY_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json',
                                  side_effect=lambda url: GOOGLE_PAYLOAD if 'googleapis' in url else {}):
            data = lookup_providers.get_provider('auto').fetch('9780132350884')
        self.assertEqual(data['title'], 'Clean Code')

    def test_cover_download_is_restricted_to_provider_hosts(self):
        with mock.patch.object(lookup_providers, '_http_get_bytes') as m:
            self.assertIsNone(lookup_providers.fetch_cover('https://evil.example/cover.jpg'))
            self.assertIsNone(lookup_providers.fetch_cover('file:///etc/passwd'))
            m.assert_not_called()
        with mock.patch.object(lookup_providers, '_http_get_bytes',
                               return_value=(b'IMG', 'image/jpeg')):
            self.assertEqual(lookup_providers.fetch_cover(
                'https://portal.dnb.de/opac/mvb/cover?isbn=1'), (b'IMG', 'jpg'))
        with mock.patch.object(lookup_providers, '_http_get_bytes',
                               return_value=(b'<html>', 'text/html')):
            self.assertIsNone(lookup_providers.fetch_cover('https://portal.dnb.de/opac/mvb/cover?isbn=1'))

    def test_dnb_search_builds_word_query_and_parses_records(self):
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_SEARCH_PAYLOAD) as m:
            results = lookup_providers.get_provider('dnb').search('kurze antworten hawking')
        url = m.call_args[0][0]
        self.assertIn('WOE%3Dkurze+and+WOE%3Dantworten+and+WOE%3Dhawking', url)
        self.assertEqual(len(results), 2)
        self.assertEqual(results[0]['title'], 'Kurze Antworten auf große Fragen')
        self.assertEqual(results[0]['isbn'], '9783608963762')
        self.assertEqual(results[1]['year'], '1988')

    def test_openlibrary_search_normalises_docs(self):
        with mock.patch.object(lookup_providers, '_http_get_json', return_value=OPENLIBRARY_SEARCH_PAYLOAD):
            results = lookup_providers.get_provider('openlibrary').search('clean code')
        self.assertEqual(len(results), 2)
        self.assertEqual(results[0]['title'], 'Clean Code')
        self.assertEqual(results[0]['isbn'], '9780132350884')  # ISBN-13 preferred
        self.assertIn('covers.openlibrary.org/b/id/123', results[0]['cover_url'])
        self.assertNotIn('isbn', results[1])  # missing data stays absent

    def test_google_search_returns_candidates(self):
        with mock.patch.object(lookup_providers, '_http_get_json', return_value=GOOGLE_PAYLOAD):
            results = lookup_providers.get_provider('googlebooks').search('clean code')
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['title'], 'Clean Code')

    def test_auto_search_falls_back_to_next_source(self):
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_EMPTY_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json',
                                  side_effect=lambda url: GOOGLE_PAYLOAD if 'googleapis' in url else {}):
            results = lookup_providers.get_provider('auto').search('clean code')
        self.assertEqual(results[0]['title'], 'Clean Code')

    def test_auto_search_merges_all_sources_and_dedupes_by_isbn(self):
        # Every source answers: DNB contributes two records, Google and Open
        # Library both return "Clean Code" with the same ISBN — the duplicate
        # must collapse into a single entry.
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_SEARCH_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json',
                                  side_effect=lambda url: GOOGLE_PAYLOAD if 'googleapis' in url
                                  else OPENLIBRARY_SEARCH_PAYLOAD):
            results = lookup_providers.get_provider('auto').search('hawking clean code')
        titles = [r['title'] for r in results]
        self.assertIn('Kurze Antworten auf große Fragen', titles)  # DNB
        self.assertIn('The Clean Coder', titles)  # Open Library only
        self.assertEqual(titles.count('Clean Code'), 1)  # Google + Open Library deduped

    def test_no_match_returns_empty(self):
        with mock.patch.object(lookup_providers, '_http_get_json', return_value={}):
            self.assertEqual(lookup_providers.get_provider('openlibrary').fetch('000'), {})

    def test_network_failure_returns_none_then_empty(self):
        with mock.patch.object(lookup_providers, '_http_get_json', return_value=None):
            self.assertEqual(lookup_providers.get_provider('googlebooks').fetch('9780132350884'), {})


class LookupViewTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.coll = Collection.objects.create(owner=self.owner, name='Bücher', lookup_provider='openlibrary')
        add_field(self.coll, 'titel', 'Titel', FieldType.TEXT, order=0, config={'lookup_attribute': 'title'})
        add_field(self.coll, 'autor', 'Autor', FieldType.TEXT, order=1, config={'lookup_attribute': 'authors'})
        add_field(self.coll, 'isbn', 'ISBN', FieldType.ISBN, order=2, config={'lookup_attribute': 'isbn'})
        add_field(self.coll, 'notiz', 'Notiz', FieldType.TEXT, order=3)  # unmapped → never filled
        self.url = reverse('item_lookup', args=[self.coll.pk])

    def test_lookup_maps_provider_data_to_field_keys(self):
        self.client.force_login(self.owner)
        # The auto chain queries all sources: DNB (XML) answers empty here,
        # Google gets no 'items' from this payload, Open Library matches.
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_EMPTY_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json', return_value=OPENLIBRARY_PAYLOAD):
            resp = self.client.get(self.url, {'q': '9780132350884'})
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data['ok'] and data['found'])
        self.assertEqual(data['fields']['titel'], 'Clean Code')
        self.assertEqual(data['fields']['autor'], 'Robert C. Martin')
        self.assertEqual(data['fields']['isbn'], '9780132350884')
        self.assertNotIn('notiz', data['fields'])  # unmapped field untouched

    def test_lookup_requires_field_mapping(self):
        # Without any field mapped to an external attribute there is nothing
        # to fill — the endpoint must answer 400 instead of querying sources.
        for field in self.coll.fields.all():
            field.config.pop('lookup_attribute', None)
            field.save()
        self.client.force_login(self.owner)
        resp = self.client.get(self.url, {'q': '123'})
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.json()['ok'])

    def test_lookup_requires_query(self):
        self.client.force_login(self.owner)
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 400)

    def test_lookup_needs_edit_permission(self):
        CollectionShare.objects.create(collection=self.coll, user=self.viewer, permission='view')
        self.client.force_login(self.viewer)
        resp = self.client.get(self.url, {'q': '9780132350884'})
        self.assertEqual(resp.status_code, 403)

    def test_item_form_exposes_autofill_hook(self):
        self.client.force_login(self.owner)
        resp = self.client.get(reverse('item_create', args=[self.coll.pk]))
        self.assertContains(resp, 'id="autofill"')
        self.assertContains(resp, 'data-query-key="isbn"')

    def test_text_query_field_gets_scan_attribute(self):
        # Even a plain TEXT field mapped to the provider's query attribute must
        # offer camera scanning (scanner.js hooks onto data-scan).
        coll = Collection.objects.create(owner=self.owner, name='B2', lookup_provider='openlibrary')
        add_field(coll, 'code', 'Code', FieldType.TEXT, order=0, config={'lookup_attribute': 'isbn'})
        self.client.force_login(self.owner)
        resp = self.client.get(reverse('item_create', args=[coll.pk]))
        self.assertContains(resp, 'data-scan="isbn"')

    def test_lookup_flags_duplicate_isbn_but_not_the_item_itself(self):
        item = Item.objects.create(collection=self.coll,
                                   values={'name': 'Vorhanden', 'isbn': '978-0-13-235088-4'})
        self.client.force_login(self.owner)
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_EMPTY_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json', return_value=OPENLIBRARY_PAYLOAD):
            resp = self.client.get(self.url, {'q': '9780132350884'})
        duplicate = resp.json()['duplicate']
        self.assertEqual(duplicate['name'], 'Vorhanden')  # hyphens vs. digits normalised
        self.assertIn(str(item.pk), duplicate['url'])
        # Editing that very item must not warn about itself.
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_EMPTY_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json', return_value=OPENLIBRARY_PAYLOAD):
            resp = self.client.get(self.url, {'q': '9780132350884', 'exclude': str(item.pk)})
        self.assertIsNone(resp.json()['duplicate'])


class SearchViewTests(TestCase):
    """Free-text search: suggest external records and link one to the item."""

    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.coll = Collection.objects.create(owner=self.owner, name='Bücher',
                                              lookup_provider='openlibrary')
        add_field(self.coll, 'titel', 'Titel', FieldType.TEXT, order=0, config={'lookup_attribute': 'title'})
        add_field(self.coll, 'autor', 'Autor', FieldType.TEXT, order=1, config={'lookup_attribute': 'authors'})
        add_field(self.coll, 'isbn', 'ISBN', FieldType.ISBN, order=2, config={'lookup_attribute': 'isbn'})
        self.url = reverse('item_search', args=[self.coll.pk])
        self.client.force_login(self.owner)

    def test_search_maps_candidates_to_field_keys(self):
        # DNB answers empty, Google finds no 'items' in this payload — only
        # Open Library's two candidates survive the merged auto search.
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_EMPTY_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json',
                                  return_value=OPENLIBRARY_SEARCH_PAYLOAD):
            resp = self.client.get(self.url, {'q': 'clean code martin'})
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data['ok'])
        self.assertEqual(len(data['results']), 2)
        first = data['results'][0]
        self.assertEqual(first['fields']['titel'], 'Clean Code')
        self.assertEqual(first['fields']['isbn'], '9780132350884')  # link = ISBN filled too
        self.assertIn('Clean Code', first['label'])
        self.assertIn('Robert C. Martin', first['label'])
        self.assertIn('covers.openlibrary.org', first['cover'])

    def test_search_requires_query_and_field_mapping(self):
        self.assertEqual(self.client.get(self.url).status_code, 400)
        for field in self.coll.fields.all():
            field.config.pop('lookup_attribute', None)
            field.save()
        self.assertEqual(self.client.get(self.url, {'q': 'x'}).status_code, 400)

    def test_search_needs_edit_permission(self):
        CollectionShare.objects.create(collection=self.coll, user=self.viewer, permission='view')
        self.client.force_login(self.viewer)
        self.assertEqual(self.client.get(self.url, {'q': 'x'}).status_code, 403)

    def test_item_form_exposes_search_hook(self):
        resp = self.client.get(reverse('item_create', args=[self.coll.pk]))
        self.assertContains(resp, 'data-search-url="%s"' % self.url)


class FindTests(TestCase):
    """Scan-to-find: a scanned code opens the matching item."""

    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Bücher')
        add_field(self.col, 'isbn', 'ISBN', FieldType.ISBN, order=0)
        self.item = Item.objects.create(collection=self.col,
                                        values={'name': 'Buch', 'isbn': '978-3-16-148410-0'})
        self.client.force_login(self.owner)
        self.url = reverse('item_find', args=[self.col.pk])
        self.detail = reverse('item_detail', args=[self.col.pk, self.item.pk])

    def test_find_by_label_barcode_short_code(self):
        resp = self.client.get(self.url, {'code': self.item.id.hex[:12]})
        self.assertRedirects(resp, self.detail)

    def test_find_by_item_qr_url(self):
        resp = self.client.get(self.url, {'code': 'http://testserver' + self.detail})
        self.assertRedirects(resp, self.detail)

    def test_find_by_isbn_value_normalises_hyphens(self):
        resp = self.client.get(self.url, {'code': '9783161484100'})
        self.assertRedirects(resp, self.detail)

    def test_unknown_code_shows_warning(self):
        resp = self.client.get(self.url, {'code': 'gibtsnicht'}, follow=True)
        self.assertContains(resp, 'Kein Gegenstand mit Code')

    def test_detail_page_has_scan_button(self):
        resp = self.client.get(reverse('collection_detail', args=[self.col.pk]))
        self.assertContains(resp, 'id="scanFind"')
        self.assertContains(resp, 'js/scanner.js')

    def test_find_requires_access(self):
        stranger = User.objects.create_user('fremd', 'f@e.de', 'pw')
        self.client.force_login(stranger)
        resp = self.client.get(self.url, {'code': 'x'})
        self.assertEqual(resp.status_code, 403)


class PresetTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.client.force_login(self.owner)

    def test_create_with_book_preset_maps_lookup_fields(self):
        self.client.post(reverse('collection_create'),
                         {'name': 'Meine Bücher', 'description': '', 'preset': 'books'})
        coll = Collection.objects.get(name='Meine Bücher')
        isbn = coll.fields.get(key='isbn')
        self.assertEqual(isbn.config.get('lookup_attribute'), 'isbn')
        self.assertEqual(coll.fields.get(key='titel').config.get('lookup_attribute'), 'title')

    def test_non_book_presets_create_their_fields(self):
        from .services import PRESETS
        for preset in ('movies', 'music', 'games'):
            name = 'Vorlage %s' % preset
            self.client.post(reverse('collection_create'),
                             {'name': name, 'description': '', 'preset': preset})
            coll = Collection.objects.get(name=name)
            expected = [spec[0] for spec in PRESETS[preset]['fields']]
            self.assertEqual(
                list(coll.fields.order_by('order').values_list('key', flat=True)),
                expected, preset)

    def test_preset_choices_offered_in_form(self):
        resp = self.client.get(reverse('collection_create'))
        for value in ('books', 'movies', 'music', 'games'):
            self.assertContains(resp, 'value="%s"' % value)


# --- Internationalisation (German default + English) --------------------------

class I18nTests(TestCase):
    def test_catalogue_translates_strings_to_english(self):
        from django.utils import translation
        with translation.override('en'):
            self.assertEqual(translation.gettext('Sammlungen'), 'Collections')
            self.assertEqual(translation.gettext('Speichern'), 'Save')
        with translation.override('de'):
            # German is the source language: msgid is returned unchanged.
            self.assertEqual(translation.gettext('Sammlungen'), 'Sammlungen')

    def test_login_page_renders_in_german_by_default(self):
        resp = self.client.get(reverse('login'))
        self.assertContains(resp, 'Anmelden')

    def test_english_browser_header_is_ignored_without_explicit_choice(self):
        # DefaultLanguageMiddleware: an English Accept-Language header alone
        # must NOT switch the UI — German stays the default.
        resp = self.client.get(reverse('login'), HTTP_ACCEPT_LANGUAGE='en-US,en;q=0.9')
        self.assertContains(resp, 'Anmelden')
        self.assertNotContains(resp, 'Sign in')

    def test_explicit_language_cookie_beats_default(self):
        from django.conf import settings as dj_settings
        self.client.cookies[dj_settings.LANGUAGE_COOKIE_NAME] = 'en'
        resp = self.client.get(reverse('login'), HTTP_ACCEPT_LANGUAGE='en-US,en;q=0.9')
        self.assertContains(resp, 'Sign in')

    def test_set_language_switches_ui_to_english(self):
        # Persist English via the set_language endpoint, then a page renders EN.
        resp = self.client.post(reverse('set_language'),
                                {'language': 'en', 'next': reverse('login')},
                                follow=True)
        self.assertContains(resp, 'Sign in')
        self.assertNotContains(resp, 'Registrieren')

    def test_language_switcher_present_in_page(self):
        resp = self.client.get(reverse('login'))
        self.assertContains(resp, reverse('set_language'))

    def test_javascript_catalogue_served_and_translated(self):
        # Default (German = source language): the catalogue is served but
        # carries no translations. With English chosen, the JS strings appear.
        url = reverse('javascript-catalog')
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn('javascript', resp['Content-Type'])
        from django.conf import settings as dj_settings
        self.client.cookies[dj_settings.LANGUAGE_COOKIE_NAME] = 'en'
        resp = self.client.get(url)
        self.assertContains(resp, 'Search suggestions')  # 'Vorschläge suchen'

    def test_base_template_loads_js_catalogue(self):
        resp = self.client.get(reverse('login'))
        self.assertContains(resp, reverse('javascript-catalog'))


class LoanTests(TestCase):
    """Lending: mark items as lent, return them, list open loans."""

    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Bücher')
        self.item = Item.objects.create(collection=self.col, values={'name': 'Dune'})
        CollectionShare.objects.create(collection=self.col, user=self.viewer, permission='view')
        self.client.force_login(self.owner)
        self.lend_url = reverse('item_lend', args=[self.col.pk, self.item.pk])
        self.return_url = reverse('item_return', args=[self.col.pk, self.item.pk])

    def test_lend_and_return(self):
        resp = self.client.post(self.lend_url, {'borrower': 'Anna', 'note': 'bis August'})
        self.assertEqual(resp.status_code, 302)
        loan = self.item.active_loan
        self.assertEqual(loan.borrower, 'Anna')
        self.assertIsNone(loan.returned_at)
        self.client.post(self.return_url)
        loan.refresh_from_db()
        self.assertIsNotNone(loan.returned_at)
        self.assertIsNone(self.item.active_loan)

    def test_double_lend_blocked(self):
        self.client.post(self.lend_url, {'borrower': 'Anna'})
        self.client.post(self.lend_url, {'borrower': 'Bernd'})
        self.assertEqual(self.item.loans.count(), 1)

    def test_borrower_required(self):
        self.client.post(self.lend_url, {'borrower': '  '})
        self.assertEqual(self.item.loans.count(), 0)

    def test_viewer_cannot_lend_but_can_see_list(self):
        self.client.force_login(self.viewer)
        resp = self.client.post(self.lend_url, {'borrower': 'X'})
        self.assertEqual(resp.status_code, 403)
        Loan.objects.create(item=self.item, borrower='Anna')
        resp = self.client.get(reverse('collection_loans', args=[self.col.pk]))
        self.assertContains(resp, 'Anna')
        self.assertNotContains(resp, 'Rückgabe vermerken')

    def test_item_detail_shows_loan_state(self):
        Loan.objects.create(item=self.item, borrower='Anna')
        resp = self.client.get(reverse('item_detail', args=[self.col.pk, self.item.pk]))
        self.assertContains(resp, 'Verliehen an')
        self.assertContains(resp, 'Anna')

    def test_detail_toolbar_shows_open_loan_badge(self):
        Loan.objects.create(item=self.item, borrower='Anna')
        resp = self.client.get(reverse('collection_detail', args=[self.col.pk]))
        self.assertContains(resp, 'Ausleihen')
        self.assertContains(resp, 'text-bg-warning">1</span>')

    def test_dashboard_lists_open_loans_and_flags_overdue(self):
        from datetime import timedelta
        from django.utils import timezone
        Loan.objects.create(item=self.item, borrower='Anna',
                            lent_at=timezone.localdate() - timedelta(days=40))
        resp = self.client.get(reverse('dashboard'))
        self.assertContains(resp, 'Offene Ausleihen')
        self.assertContains(resp, 'Anna')
        self.assertContains(resp, 'überfällig')
        # returned loans disappear from the dashboard
        Loan.objects.update(returned_at=timezone.localdate())
        resp = self.client.get(reverse('dashboard'))
        self.assertNotContains(resp, 'Offene Ausleihen')


class ImportTests(TestCase):
    """Excel/CSV import: headers map to field labels, 'Art' creates ItemTypes."""

    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Filme')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        add_field(self.col, 'preis', 'Preis', FieldType.PRICE, order=1)
        add_field(self.col, 'gesehen', 'Gesehen', FieldType.BOOLEAN, order=2)
        add_field(self.col, 'kaufdatum', 'Kaufdatum', FieldType.DATE, order=3)
        add_field(self.col, 'genre', 'Genre', FieldType.CHOICE, order=4,
                  config={'choices': ['Action', 'Drama']})
        CollectionShare.objects.create(collection=self.col, user=self.viewer, permission='view')
        self.client.force_login(self.owner)
        self.url = reverse('collection_import', args=[self.col.pk])

    def _xlsx(self, rows) -> SimpleUploadedFile:
        from openpyxl import Workbook
        wb = Workbook()
        for row in rows:
            wb.active.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return SimpleUploadedFile('daten.xlsx', buf.getvalue())

    def test_import_xlsx_creates_items_and_art(self):
        upload = self._xlsx([
            ['Art', 'Name', 'Preis', 'Gesehen', 'Kaufdatum', 'Genre', 'Unbekannt'],
            ['DVD', 'Matrix', '9,99', 'Ja', '01.05.2024', 'action', 'x'],
            ['Blu-ray', 'Dune', 19.5, 'Nein', '2024-06-01', 'Drama', ''],
        ])
        resp = self.client.post(self.url, {'file': upload})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(self.col.items.count(), 2)
        matrix = self.col.items.get(values__name='Matrix')
        self.assertEqual(matrix.item_type.name, 'DVD')
        self.assertEqual(matrix.values['preis'], 9.99)
        self.assertIs(matrix.values['gesehen'], True)
        self.assertEqual(matrix.values['kaufdatum'], '2024-05-01')
        self.assertEqual(matrix.values['genre'], 'Action')  # case-insensitive match
        self.assertContains(resp, 'Unbekannt')  # reported as ignored column

    def test_import_csv_with_semicolon(self):
        upload = SimpleUploadedFile(
            'daten.csv', 'Name;Preis\nMatrix;9,99\nDune;19.5\n'.encode('utf-8-sig'))
        resp = self.client.post(self.url, {'file': upload})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(self.col.items.count(), 2)

    def test_invalid_cells_are_skipped_with_warning(self):
        result = imports.import_table(self.col, [
            ['Name', 'Preis'],
            ['Matrix', 'teuer'],
        ])
        self.assertEqual(result['created'], 1)  # row imported, bad cell dropped
        self.assertEqual(len(result['warnings']), 1)
        self.assertNotIn('preis', self.col.items.get().values)

    def test_import_requires_edit_permission(self):
        self.client.force_login(self.viewer)
        self.assertEqual(self.client.get(self.url).status_code, 403)

    def test_wrong_file_rejected(self):
        resp = self.client.post(self.url, {'file': SimpleUploadedFile('x.txt', b'nope')},
                                follow=True)
        self.assertContains(resp, '.xlsx- oder .csv-Datei')
        self.assertEqual(self.col.items.count(), 0)


class PwaTests(TestCase):
    def test_service_worker_served_from_root(self):
        resp = self.client.get('/sw.js')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/javascript')
        self.assertContains(resp, "caches.open")

    def test_manifest_served(self):
        resp = self.client.get('/manifest.webmanifest')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/manifest+json')
        data = json.loads(resp.content)
        self.assertEqual(data['short_name'], 'CMS')
        self.assertEqual(len(data['icons']), 2)

    def test_base_template_links_manifest(self):
        resp = self.client.get(reverse('login'))
        self.assertContains(resp, 'rel="manifest"')


class CollectionEditDeleteTests(TestCase):
    """Renaming/describing and deleting whole collections."""

    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.editor = User.objects.create_user('editor', 'e@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Alt', description='x')
        CollectionShare.objects.create(collection=self.col, user=self.editor, permission='edit')
        CollectionShare.objects.create(collection=self.col, user=self.viewer, permission='view')
        self.edit_url = reverse('collection_edit', args=[self.col.pk])
        self.delete_url = reverse('collection_delete', args=[self.col.pk])

    def test_owner_can_rename(self):
        self.client.force_login(self.owner)
        resp = self.client.post(self.edit_url, {'name': 'Neu', 'description': 'y'})
        self.assertRedirects(resp, reverse('collection_detail', args=[self.col.pk]))
        self.col.refresh_from_db()
        self.assertEqual(self.col.name, 'Neu')
        self.assertEqual(self.col.description, 'y')

    def test_edit_form_hides_preset_and_template(self):
        self.client.force_login(self.owner)
        resp = self.client.get(self.edit_url)
        self.assertNotContains(resp, 'name="preset"')
        self.assertNotContains(resp, 'name="template"')

    def test_editor_can_rename_viewer_cannot(self):
        self.client.force_login(self.editor)
        self.client.post(self.edit_url, {'name': 'Vom Editor', 'description': ''})
        self.col.refresh_from_db()
        self.assertEqual(self.col.name, 'Vom Editor')
        self.client.force_login(self.viewer)
        self.assertEqual(self.client.get(self.edit_url).status_code, 403)

    def test_only_owner_can_delete(self):
        self.client.force_login(self.editor)
        self.assertEqual(self.client.post(self.delete_url).status_code, 403)
        self.client.force_login(self.owner)
        resp = self.client.post(self.delete_url)
        self.assertRedirects(resp, reverse('dashboard'))
        self.assertFalse(Collection.objects.filter(pk=self.col.pk).exists())

    def test_delete_shows_confirmation_first(self):
        self.client.force_login(self.owner)
        resp = self.client.get(self.delete_url)
        self.assertContains(resp, 'löschen')
        self.assertTrue(Collection.objects.filter(pk=self.col.pk).exists())


class SortPaginationTests(TestCase):
    """Sortable columns and pagination of the item table."""

    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Bücher')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        self.client.force_login(self.owner)
        self.url = reverse('collection_detail', args=[self.col.pk])

    def test_sort_by_field_asc_and_desc(self):
        for name in ('Charlie', 'Alpha', 'Bravo'):
            Item.objects.create(collection=self.col, values={'name': name})
        resp = self.client.get(self.url, {'sort': 'name', 'dir': 'asc'})
        names = [r['item'].values['name'] for r in resp.context['rows']]
        self.assertEqual(names, ['Alpha', 'Bravo', 'Charlie'])
        resp = self.client.get(self.url, {'sort': 'name', 'dir': 'desc'})
        names = [r['item'].values['name'] for r in resp.context['rows']]
        self.assertEqual(names, ['Charlie', 'Bravo', 'Alpha'])

    def test_unknown_sort_key_is_ignored(self):
        Item.objects.create(collection=self.col, values={'name': 'A'})
        resp = self.client.get(self.url, {'sort': 'nope"; drop', 'dir': 'desc'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context['rows']), 1)

    def test_items_are_paginated(self):
        from .runtime_settings import get_setting
        per_page = get_setting('items_per_page')
        for i in range(per_page + 5):
            Item.objects.create(collection=self.col, values={'name': f'Item {i:03d}'})
        resp = self.client.get(self.url)
        self.assertEqual(len(resp.context['rows']), per_page)
        self.assertContains(resp, 'pagination')
        resp = self.client.get(self.url, {'page': 2})
        self.assertEqual(len(resp.context['rows']), 5)

    def test_page_size_is_configurable_at_runtime(self):
        from django.core.cache import cache
        from .runtime_settings import _CACHE_KEY, set_setting
        # The settings cache outlives the test transaction: clean it up so the
        # rolled-back override can't leak into other tests.
        self.addCleanup(cache.delete, _CACHE_KEY)
        set_setting('items_per_page', 5)
        for i in range(7):
            Item.objects.create(collection=self.col, values={'name': f'Item {i:03d}'})
        resp = self.client.get(self.url)
        self.assertEqual(len(resp.context['rows']), 5)

    def test_pagination_keeps_filters(self):
        for i in range(60):
            Item.objects.create(collection=self.col, values={'name': f'Buch {i:03d}'})
        resp = self.client.get(self.url, {'q': 'Buch', 'page': 2})
        self.assertEqual(resp.context['result_count'], 60)
        self.assertEqual(len(resp.context['rows']), 10)


@override_settings(MEDIA_ROOT=MEDIA)
class ItemDuplicateTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Bücher')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        add_field(self.col, 'bild', 'Bild', FieldType.IMAGE, order=1)
        CollectionShare.objects.create(collection=self.col, user=self.viewer, permission='view')
        self.client.force_login(self.owner)

    def test_duplicate_copies_values_type_and_assets(self):
        art = ItemType.objects.create(collection=self.col, name='Roman')
        resp = self.client.post(reverse('item_create', args=[self.col.pk]), {
            '__item_type': art.pk, 'name': 'Dune',
            'bild': SimpleUploadedFile('cover.png', make_png(), content_type='image/png'),
        })
        self.assertEqual(resp.status_code, 302)
        original = self.col.items.get()
        resp = self.client.post(reverse('item_duplicate', args=[self.col.pk, original.pk]))
        self.assertEqual(self.col.items.count(), 2)
        copy = self.col.items.exclude(pk=original.pk).get()
        self.assertRedirects(resp, reverse('item_edit', args=[self.col.pk, copy.pk]))
        self.assertEqual(copy.values['name'], 'Dune')
        self.assertEqual(copy.item_type, art)
        self.assertEqual(copy.assets.count(), 1)
        # The copy owns its own file — not a reference to the original's asset.
        self.assertNotEqual(copy.assets.get().file.name, original.assets.get().file.name)
        self.assertEqual(copy.values['bild']['asset_id'], str(copy.assets.get().id))

    def test_viewer_cannot_duplicate(self):
        item = Item.objects.create(collection=self.col, values={'name': 'Dune'})
        self.client.force_login(self.viewer)
        resp = self.client.post(reverse('item_duplicate', args=[self.col.pk, item.pk]))
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(self.col.items.count(), 1)


class LoanDueDateTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Bücher')
        self.item = Item.objects.create(collection=self.col, values={'name': 'Dune'})
        self.client.force_login(self.owner)
        self.lend_url = reverse('item_lend', args=[self.col.pk, self.item.pk])

    def test_lend_with_due_date(self):
        self.client.post(self.lend_url, {'borrower': 'Anna', 'due_at': '2026-08-01'})
        loan = self.item.active_loan
        self.assertEqual(str(loan.due_at), '2026-08-01')

    def test_invalid_due_date_is_ignored(self):
        self.client.post(self.lend_url, {'borrower': 'Anna', 'due_at': 'kein-datum'})
        loan = self.item.active_loan
        self.assertIsNotNone(loan)
        self.assertIsNone(loan.due_at)

    def test_overdue_uses_due_date(self):
        from datetime import timedelta
        from django.utils import timezone
        today = timezone.localdate()
        loan = Loan.objects.create(item=self.item, borrower='Anna',
                                   due_at=today - timedelta(days=1))
        self.assertTrue(loan.is_overdue)
        loan.due_at = today + timedelta(days=1)
        self.assertFalse(loan.is_overdue)
        # Returned loans are never overdue.
        loan.due_at = today - timedelta(days=1)
        loan.returned_at = today
        self.assertFalse(loan.is_overdue)


class GlobalSearchTests(TestCase):
    """The navbar search spans all accessible collections and their items."""

    def setUp(self):
        self.user = User.objects.create_user('anna', 'a@e.de', 'pw')
        self.other = User.objects.create_user('bernd', 'b@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.user, name='Science-Fiction',
                                             description='Romane und Filme')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        self.item = Item.objects.create(collection=self.col, values={'name': 'Dune'})
        # A foreign collection that must never appear in the results.
        foreign = Collection.objects.create(owner=self.other, name='Dune-Sammlung')
        add_field(foreign, 'name', 'Name', FieldType.TEXT, order=0)
        Item.objects.create(collection=foreign, values={'name': 'Dune Geheim'})
        self.client.force_login(self.user)
        self.url = reverse('global_search')

    def test_requires_login(self):
        self.client.logout()
        resp = self.client.get(self.url, {'q': 'Dune'})
        self.assertEqual(resp.status_code, 302)

    def test_finds_collections_by_name_and_description(self):
        resp = self.client.get(self.url, {'q': 'science'})
        self.assertEqual([c.pk for c in resp.context['collection_hits']], [self.col.pk])
        resp = self.client.get(self.url, {'q': 'Romane'})
        self.assertEqual([c.pk for c in resp.context['collection_hits']], [self.col.pk])

    def test_finds_items_only_in_accessible_collections(self):
        resp = self.client.get(self.url, {'q': 'dune'})
        self.assertEqual([i.pk for i in resp.context['item_hits']], [self.item.pk])
        self.assertNotContains(resp, 'Dune Geheim')

    def test_finds_items_by_type_name(self):
        art = ItemType.objects.create(collection=self.col, name='Taschenbuch')
        self.item.item_type = art
        self.item.save()
        resp = self.client.get(self.url, {'q': 'taschenbuch'})
        self.assertEqual([i.pk for i in resp.context['item_hits']], [self.item.pk])

    def test_shared_collections_are_searched(self):
        self.client.force_login(self.other)
        CollectionShare.objects.create(collection=self.col, user=self.other, permission='view')
        resp = self.client.get(self.url, {'q': 'dune'})
        self.assertIn(self.item.pk, [i.pk for i in resp.context['item_hits']])

    def test_empty_query_shows_hint(self):
        resp = self.client.get(self.url)
        self.assertContains(resp, 'Suchbegriff')

    def test_navbar_contains_search_form(self):
        resp = self.client.get(reverse('dashboard'))
        self.assertContains(resp, reverse('global_search'))


class BulkActionTests(TestCase):
    """Bulk delete / bulk "Art" assignment from the item table."""

    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Bücher')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        CollectionShare.objects.create(collection=self.col, user=self.viewer, permission='view')
        self.items = [Item.objects.create(collection=self.col, values={'name': f'Buch {i}'})
                      for i in range(3)]
        self.client.force_login(self.owner)
        self.url = reverse('items_bulk', args=[self.col.pk])

    def test_bulk_delete(self):
        resp = self.client.post(self.url, {
            'action': 'delete',
            'items': [str(self.items[0].pk), str(self.items[1].pk)],
        })
        self.assertRedirects(resp, reverse('collection_detail', args=[self.col.pk]))
        self.assertEqual(self.col.items.count(), 1)

    def test_bulk_set_and_clear_type(self):
        art = ItemType.objects.create(collection=self.col, name='Roman')
        self.client.post(self.url, {
            'action': 'set_type', 'item_type': str(art.pk),
            'items': [str(i.pk) for i in self.items],
        })
        self.assertEqual(self.col.items.filter(item_type=art).count(), 3)
        self.client.post(self.url, {
            'action': 'set_type', 'item_type': '',
            'items': [str(self.items[0].pk)],
        })
        self.assertEqual(self.col.items.filter(item_type=art).count(), 2)

    def test_foreign_items_and_invalid_ids_are_ignored(self):
        foreign_col = Collection.objects.create(owner=self.viewer, name='Fremd')
        foreign = Item.objects.create(collection=foreign_col, values={})
        self.client.post(self.url, {
            'action': 'delete',
            'items': [str(foreign.pk), 'kein-uuid'],
        })
        self.assertTrue(Collection.objects.get(pk=foreign_col.pk).items.filter(pk=foreign.pk).exists())
        self.assertEqual(self.col.items.count(), 3)

    def test_type_from_other_collection_rejected(self):
        foreign_col = Collection.objects.create(owner=self.owner, name='Andere')
        foreign_art = ItemType.objects.create(collection=foreign_col, name='Fremd')
        resp = self.client.post(self.url, {
            'action': 'set_type', 'item_type': str(foreign_art.pk),
            'items': [str(self.items[0].pk)],
        })
        self.assertEqual(resp.status_code, 404)
        self.assertIsNone(self.col.items.get(pk=self.items[0].pk).item_type)

    def test_viewer_cannot_bulk_edit(self):
        self.client.force_login(self.viewer)
        resp = self.client.post(self.url, {'action': 'delete',
                                           'items': [str(self.items[0].pk)]})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(self.col.items.count(), 3)

    def test_checkboxes_only_rendered_for_editors(self):
        resp = self.client.get(reverse('collection_detail', args=[self.col.pk]))
        self.assertContains(resp, 'data-bulk-item')
        self.client.force_login(self.viewer)
        resp = self.client.get(reverse('collection_detail', args=[self.col.pk]))
        self.assertNotContains(resp, 'data-bulk-item')


class RuntimeSettingsTests(TestCase):
    """Database-backed runtime settings: precedence, validation, caching."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        self.rs = runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)

    def test_code_default_without_any_override(self):
        self.assertEqual(self.rs.get_setting('items_per_page'), 50)
        self.assertFalse(self.rs.get_setting('registration_auto_approve'))

    def test_db_override_wins(self):
        self.rs.set_setting('items_per_page', 25)
        self.assertEqual(self.rs.get_setting('items_per_page'), 25)

    def test_ini_fallback_used_when_no_db_row(self):
        with mock.patch.dict('cms.conf._INI_VALUES', {'ITEMS_PER_PAGE': '75'}):
            self.assertEqual(self.rs.get_setting('items_per_page'), 75)

    def test_db_override_beats_ini(self):
        self.rs.set_setting('items_per_page', 25)
        with mock.patch.dict('cms.conf._INI_VALUES', {'ITEMS_PER_PAGE': '75'}):
            self.assertEqual(self.rs.get_setting('items_per_page'), 25)

    def test_set_setting_validates(self):
        with self.assertRaises(ValueError):
            self.rs.set_setting('items_per_page', 1)  # below min_value=5
        with self.assertRaises(ValueError):
            self.rs.set_setting('items_per_page', 'keine Zahl')

    def test_bool_and_str_coercion(self):
        self.rs.set_setting('registration_auto_approve', 'true')
        self.assertIs(self.rs.get_setting('registration_auto_approve'), True)
        self.rs.set_setting('default_currency', 'USD')
        self.assertEqual(self.rs.get_setting('default_currency'), 'USD')

    def test_corrupt_db_value_falls_back_to_default(self):
        from .models import SiteSetting
        SiteSetting.objects.create(key='items_per_page', value='kaputt')
        self.assertEqual(self.rs.get_setting('items_per_page'), 50)

    def test_default_currency_applies_to_new_collections(self):
        self.rs.set_setting('default_currency', 'CHF')
        owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        col = Collection.objects.create(owner=owner, name='Münzen')
        create_default_fields(col)
        preis = col.fields.get(key='preis')
        self.assertEqual(preis.config.get('currency'), 'CHF')

    def test_loan_overdue_days_configurable(self):
        from datetime import timedelta
        from django.utils import timezone
        owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        col = Collection.objects.create(owner=owner, name='Bücher')
        item = Item.objects.create(collection=col, values={'name': 'Dune'})
        loan = Loan.objects.create(
            item=item, borrower='Anna',
            lent_at=timezone.localdate() - timedelta(days=10),
        )
        self.assertFalse(loan.is_overdue)  # default: 30 days
        self.rs.set_setting('loan_overdue_days', 7)
        self.assertTrue(loan.is_overdue)


class SiteSettingsPageTests(TestCase):
    """The staff-only settings page (form generated from the registry)."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.staff = User.objects.create_user('chef', 'c@e.de', 'pw', is_staff=True)
        self.user = User.objects.create_user('normal', 'n@e.de', 'pw')
        self.url = reverse('site_settings')

    def test_requires_staff(self):
        self.client.force_login(self.user)
        self.assertEqual(self.client.get(self.url).status_code, 403)

    def test_anonymous_redirected_to_login(self):
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse('login'), resp['Location'])

    def test_staff_can_view_and_save(self):
        from . import runtime_settings
        self.client.force_login(self.staff)
        resp = self.client.get(self.url)
        self.assertContains(resp, 'items_per_page')

        data = {key: runtime_settings.get_setting(key)
                for key, d in runtime_settings.REGISTRY.items() if d.kind != 'bool'}
        data['items_per_page'] = 20
        data['registration_auto_approve'] = 'on'
        resp = self.client.post(self.url, data)
        self.assertRedirects(resp, self.url)
        self.assertEqual(runtime_settings.get_setting('items_per_page'), 20)
        self.assertTrue(runtime_settings.get_setting('registration_auto_approve'))

    def test_invalid_value_rejected(self):
        from . import runtime_settings
        self.client.force_login(self.staff)
        data = {key: runtime_settings.get_setting(key)
                for key, d in runtime_settings.REGISTRY.items() if d.kind != 'bool'}
        data['items_per_page'] = 100000  # above max_value
        resp = self.client.post(self.url, data)
        self.assertEqual(resp.status_code, 200)  # re-rendered with errors
        self.assertEqual(runtime_settings.get_setting('items_per_page'), 50)

    def test_nav_link_only_for_staff(self):
        self.client.force_login(self.staff)
        self.assertContains(self.client.get(reverse('dashboard')), self.url)
        self.client.force_login(self.user)
        self.assertNotContains(self.client.get(reverse('dashboard')), self.url)


@override_settings(MEDIA_ROOT=MEDIA)
class UploadLimitTests(TestCase):
    """Runtime-configurable upload limits (size and allowed extensions)."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        self.rs = runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Dokumente')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        add_field(self.col, 'beleg', 'Beleg', FieldType.FILE, order=1)
        self.client.force_login(self.owner)
        self.url = reverse('item_create', args=[self.col.pk])

    def test_oversized_file_rejected(self):
        self.rs.set_setting('upload_max_mb', 1)
        big = SimpleUploadedFile('beleg.pdf', b'x' * (1024 * 1024 + 1))
        resp = self.client.post(self.url, {'name': 'Quittung', 'beleg': big})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'zu groß')
        self.assertEqual(self.col.items.count(), 0)

    def test_file_within_limit_accepted(self):
        self.rs.set_setting('upload_max_mb', 1)
        small = SimpleUploadedFile('beleg.pdf', b'x' * 1024)
        resp = self.client.post(self.url, {'name': 'Quittung', 'beleg': small})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.col.items.count(), 1)

    def test_disallowed_extension_rejected(self):
        self.rs.set_setting('upload_allowed_extensions', 'pdf, png')
        exe = SimpleUploadedFile('virus.exe', b'MZ...')
        resp = self.client.post(self.url, {'name': 'Böse', 'beleg': exe})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'nicht erlaubt')
        self.assertEqual(self.col.items.count(), 0)

    def test_allowed_extension_accepted(self):
        self.rs.set_setting('upload_allowed_extensions', '.PDF')  # dot/case normalised
        pdf = SimpleUploadedFile('beleg.pdf', b'%PDF-1.4')
        resp = self.client.post(self.url, {'name': 'Quittung', 'beleg': pdf})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.col.items.count(), 1)

    def test_empty_extension_list_allows_everything(self):
        anything = SimpleUploadedFile('daten.xyz', b'ok')
        resp = self.client.post(self.url, {'name': 'Frei', 'beleg': anything})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.col.items.count(), 1)


class SiteSettingAuditTests(TestCase):
    """Saved settings record who changed them (``updated_by``)."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)

    def test_settings_page_records_editor(self):
        from . import runtime_settings
        from .models import SiteSetting
        staff = User.objects.create_user('chef', 'c@e.de', 'pw', is_staff=True)
        self.client.force_login(staff)
        data = {key: runtime_settings.get_setting(key)
                for key, d in runtime_settings.REGISTRY.items() if d.kind != 'bool'}
        data['registration_enabled'] = 'on'
        data['items_per_page'] = 30
        resp = self.client.post(reverse('site_settings'), data)
        self.assertRedirects(resp, reverse('site_settings'))
        row = SiteSetting.objects.get(key='items_per_page')
        self.assertEqual(row.updated_by, staff)


class AnnouncementBannerTests(TestCase):
    """Site-wide announcement banner driven by runtime settings."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        self.rs = runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.user = User.objects.create_user('u', 'u@e.de', 'pw')
        self.client.force_login(self.user)

    def test_no_banner_by_default(self):
        resp = self.client.get(reverse('dashboard'))
        self.assertNotContains(resp, 'bi-megaphone')

    def test_banner_shown_with_configured_style(self):
        self.rs.set_setting('announcement_text', 'Wartung am Sonntag!')
        self.rs.set_setting('announcement_style', 'danger')
        resp = self.client.get(reverse('dashboard'))
        self.assertContains(resp, 'Wartung am Sonntag!')
        self.assertContains(resp, 'alert-danger')

    def test_banner_also_on_login_page(self):
        self.rs.set_setting('announcement_text', 'Wartung am Sonntag!')
        self.client.logout()
        self.assertContains(self.client.get(reverse('login')), 'Wartung am Sonntag!')

    def test_invalid_style_rejected(self):
        with self.assertRaises(ValueError):
            self.rs.set_setting('announcement_style', 'evil"><script>')


class SettingsExportTests(TestCase):
    """INI export of the effective runtime settings (staff only)."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.staff = User.objects.create_user('chef', 'c@e.de', 'pw', is_staff=True)
        self.url = reverse('site_settings_export')

    def test_requires_staff(self):
        self.client.force_login(User.objects.create_user('normal', 'n@e.de', 'pw'))
        self.assertEqual(self.client.get(self.url).status_code, 403)

    def test_export_is_valid_ini_with_current_values(self):
        import configparser
        from . import runtime_settings
        runtime_settings.set_setting('items_per_page', 33)
        self.client.force_login(self.staff)
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn('attachment', resp['Content-Disposition'])
        parser = configparser.ConfigParser()
        parser.read_string(resp.content.decode())
        self.assertEqual(parser['app-defaults']['items_per_page'], '33')
        self.assertIn('registration_enabled', parser['app-defaults'])

    def test_export_round_trips_through_conf_layer(self):
        """The exported values, fed back through the INI layer, resolve identically."""
        from . import runtime_settings
        runtime_settings.set_setting('items_per_page', 44)
        runtime_settings.set_setting('registration_auto_approve', True)
        self.client.force_login(self.staff)
        body = self.client.get(self.url).content.decode()

        import configparser
        parser = configparser.ConfigParser()
        parser.read_string(body)
        ini_values = {k.upper(): v for k, v in parser.items('app-defaults')}
        from .models import SiteSetting
        SiteSetting.objects.all().delete()
        from django.core.cache import cache
        cache.delete(runtime_settings._CACHE_KEY)
        with mock.patch.dict('cms.conf._INI_VALUES', ini_values):
            self.assertEqual(runtime_settings.get_setting('items_per_page'), 44)
            self.assertIs(runtime_settings.get_setting('registration_auto_approve'), True)


class MaintenanceModeTests(TestCase):
    """Runtime-toggleable maintenance mode (middleware)."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        self.rs = runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.user = User.objects.create_user('u', 'u@e.de', 'pw')
        self.staff = User.objects.create_user('chef', 'c@e.de', 'pw', is_staff=True)

    def test_off_by_default(self):
        self.client.force_login(self.user)
        self.assertEqual(self.client.get(reverse('dashboard')).status_code, 200)

    def test_non_staff_sees_503_maintenance_page(self):
        self.rs.set_setting('maintenance_mode', True)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('dashboard'))
        self.assertEqual(resp.status_code, 503)
        self.assertContains(resp, 'Wartungsarbeiten', status_code=503)

    def test_staff_keeps_access_and_sees_notice(self):
        self.rs.set_setting('maintenance_mode', True)
        self.client.force_login(self.staff)
        resp = self.client.get(reverse('dashboard'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Wartungsmodus aktiv')

    def test_login_stays_reachable(self):
        self.rs.set_setting('maintenance_mode', True)
        self.assertEqual(self.client.get(reverse('login')).status_code, 200)

    def test_register_page_is_blocked(self):
        self.rs.set_setting('maintenance_mode', True)
        self.assertEqual(self.client.get(reverse('register')).status_code, 503)

    def test_anonymous_gets_maintenance_page(self):
        self.rs.set_setting('maintenance_mode', True)
        self.assertEqual(self.client.get(reverse('dashboard')).status_code, 503)


class SettingChangeHistoryTests(TestCase):
    """Audit history of runtime-setting changes."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        self.rs = runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.staff = User.objects.create_user('chef', 'c@e.de', 'pw', is_staff=True)

    def test_change_is_recorded_with_old_and_new_value(self):
        from .models import SettingChange
        self.rs.set_setting('items_per_page', 25, user=self.staff)
        change = SettingChange.objects.get()
        self.assertEqual((change.key, change.old_value, change.new_value),
                         ('items_per_page', 50, 25))
        self.assertEqual(change.changed_by, self.staff)

    def test_noop_save_is_not_recorded(self):
        from .models import SettingChange
        self.rs.set_setting('items_per_page', 50)  # equals the default
        self.assertEqual(SettingChange.objects.count(), 0)

    def test_settings_page_shows_history(self):
        self.rs.set_setting('items_per_page', 25, user=self.staff)
        self.client.force_login(self.staff)
        resp = self.client.get(reverse('site_settings'))
        self.assertContains(resp, 'Letzte Änderungen')
        self.assertContains(resp, '<code>50</code>', html=True)
        self.assertContains(resp, '<code>25</code>', html=True)

    def test_form_save_records_only_actual_changes(self):
        from . import runtime_settings
        from .models import SettingChange
        self.client.force_login(self.staff)
        data = {key: runtime_settings.get_setting(key)
                for key, d in runtime_settings.REGISTRY.items() if d.kind != 'bool'}
        data['registration_enabled'] = 'on'
        data['items_per_page'] = 20
        self.client.post(reverse('site_settings'), data)
        changed_keys = set(SettingChange.objects.values_list('key', flat=True))
        self.assertEqual(changed_keys, {'items_per_page'})


class UserPreferenceTests(TestCase):
    """Per-user overrides of per_user runtime settings via the profile page."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        self.rs = runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.user = User.objects.create_user('u', 'u@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.user, name='Bücher')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        self.client.force_login(self.user)

    def test_get_setting_for_prefers_user_override(self):
        self.user.preferences = {'items_per_page': 7}
        self.assertEqual(self.rs.get_setting_for(self.user, 'items_per_page'), 7)
        self.assertEqual(self.rs.get_setting('items_per_page'), 50)

    def test_non_per_user_setting_ignores_preferences(self):
        self.user.preferences = {'loan_overdue_days': 1}
        self.assertEqual(self.rs.get_setting_for(self.user, 'loan_overdue_days'), 30)

    def test_corrupt_preference_falls_back(self):
        self.user.preferences = {'items_per_page': 'kaputt'}
        self.assertEqual(self.rs.get_setting_for(self.user, 'items_per_page'), 50)

    def test_profile_saves_and_clears_override(self):
        resp = self.client.post(reverse('profile'), {
            'display_name': 'Udo', 'email': 'u@e.de', 'items_per_page': 10,
        })
        self.assertRedirects(resp, reverse('profile'))
        self.user.refresh_from_db()
        self.assertEqual(self.user.preferences, {'items_per_page': 10})
        self.assertEqual(self.user.display_name, 'Udo')

        resp = self.client.post(reverse('profile'), {
            'display_name': 'Udo', 'email': 'u@e.de', 'items_per_page': '',
        })
        self.assertRedirects(resp, reverse('profile'))
        self.user.refresh_from_db()
        self.assertEqual(self.user.preferences, {})

    def test_item_table_uses_personal_page_size(self):
        self.user.preferences = {'items_per_page': 5}
        self.user.save(update_fields=['preferences'])
        for i in range(7):
            Item.objects.create(collection=self.col, values={'name': f'Item {i}'})
        resp = self.client.get(reverse('collection_detail', args=[self.col.pk]))
        self.assertEqual(len(resp.context['rows']), 5)

    def test_other_users_unaffected_by_personal_override(self):
        self.user.preferences = {'items_per_page': 5}
        self.user.save(update_fields=['preferences'])
        other = User.objects.create_user('other', 'x@e.de', 'pw')
        CollectionShare.objects.create(collection=self.col, user=other, permission='view')
        for i in range(7):
            Item.objects.create(collection=self.col, values={'name': f'Item {i}'})
        self.client.force_login(other)
        resp = self.client.get(reverse('collection_detail', args=[self.col.pk]))
        self.assertEqual(len(resp.context['rows']), 7)

    def test_profile_rejects_out_of_range_value(self):
        resp = self.client.post(reverse('profile'), {
            'display_name': '', 'email': 'u@e.de', 'items_per_page': 1,
        })
        self.assertEqual(resp.status_code, 200)  # re-rendered with error
        self.user.refresh_from_db()
        self.assertEqual(self.user.preferences, {})


@override_settings(MEDIA_ROOT=MEDIA)
class TrashTests(TestCase):
    """Soft delete: items move to a per-collection trash and can be restored."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        self.rs = runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Bücher')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        CollectionShare.objects.create(collection=self.col, user=self.viewer, permission='view')
        self.item = Item.objects.create(collection=self.col, values={'name': 'Dune'})
        self.client.force_login(self.owner)

    def test_delete_view_moves_item_to_trash(self):
        resp = self.client.post(reverse('item_delete', args=[self.col.pk, self.item.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.col.items.count(), 0)  # default manager hides it
        self.assertEqual(Item.all_objects.filter(collection=self.col).count(), 1)
        # Trashed items are unreachable through the normal views.
        resp = self.client.get(reverse('item_detail', args=[self.col.pk, self.item.pk]))
        self.assertEqual(resp.status_code, 404)

    def test_bulk_delete_moves_to_trash(self):
        second = Item.objects.create(collection=self.col, values={'name': 'Hyperion'})
        self.client.post(reverse('items_bulk', args=[self.col.pk]), {
            'action': 'delete', 'items': [str(self.item.pk), str(second.pk)],
        })
        self.assertEqual(self.col.items.count(), 0)
        self.assertEqual(Item.all_objects.filter(collection=self.col,
                                                 deleted_at__isnull=False).count(), 2)

    def test_trash_page_lists_and_restores(self):
        self.item.soft_delete()
        resp = self.client.get(reverse('collection_trash', args=[self.col.pk]))
        self.assertContains(resp, 'Dune')
        resp = self.client.post(reverse('item_restore', args=[self.col.pk, self.item.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.col.items.count(), 1)
        self.item.refresh_from_db()
        self.assertIsNone(self.item.deleted_at)

    def test_purge_deletes_item_and_files(self):
        import os
        resp = self.client.post(reverse('item_create', args=[self.col.pk]), {'name': 'MitBild'})
        item = self.col.items.get(values__name='MitBild')
        from .models import ItemAsset
        from django.core.files.base import ContentFile
        asset = ItemAsset.objects.create(item=item, field_key='bild',
                                         file=ContentFile(make_png(), name='bild.png'))
        path = asset.file.path
        self.assertTrue(os.path.exists(path))
        item.soft_delete()
        resp = self.client.post(reverse('item_purge', args=[self.col.pk, item.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Item.all_objects.filter(pk=item.pk).exists())
        self.assertFalse(os.path.exists(path))

    def test_empty_trash(self):
        self.item.soft_delete()
        Item.objects.create(collection=self.col, values={'name': 'Bleibt'})
        self.client.post(reverse('trash_empty', args=[self.col.pk]))
        self.assertEqual(Item.all_objects.filter(collection=self.col).count(), 1)
        self.assertEqual(self.col.items.count(), 1)

    def test_retention_purges_old_items_on_trash_open(self):
        from datetime import timedelta
        from django.utils import timezone
        self.item.soft_delete()
        Item.all_objects.filter(pk=self.item.pk).update(
            deleted_at=timezone.now() - timedelta(days=31))
        self.client.get(reverse('collection_trash', args=[self.col.pk]))
        self.assertFalse(Item.all_objects.filter(pk=self.item.pk).exists())

    def test_retention_keeps_recent_items(self):
        self.item.soft_delete()
        self.client.get(reverse('collection_trash', args=[self.col.pk]))
        self.assertTrue(Item.all_objects.filter(pk=self.item.pk).exists())

    def test_viewer_cannot_access_trash_or_restore(self):
        self.item.soft_delete()
        self.client.force_login(self.viewer)
        self.assertEqual(self.client.get(
            reverse('collection_trash', args=[self.col.pk])).status_code, 403)
        self.assertEqual(self.client.post(
            reverse('item_restore', args=[self.col.pk, self.item.pk])).status_code, 403)

    def test_dashboard_count_excludes_trash(self):
        self.item.soft_delete()
        resp = self.client.get(reverse('dashboard'))
        self.assertEqual(resp.context['total_items'], 0)

    def test_purge_trash_command(self):
        from datetime import timedelta
        from django.core.management import call_command
        from django.utils import timezone
        self.item.soft_delete()
        Item.all_objects.filter(pk=self.item.pk).update(
            deleted_at=timezone.now() - timedelta(days=31))
        recent = Item.objects.create(collection=self.col, values={'name': 'Frisch'})
        recent.soft_delete()
        call_command('purge_trash', stdout=io.StringIO())
        self.assertFalse(Item.all_objects.filter(pk=self.item.pk).exists())
        self.assertTrue(Item.all_objects.filter(pk=recent.pk).exists())


class LoanReminderTests(TestCase):
    """The send_loan_reminders management command."""

    def setUp(self):
        from django.core.cache import cache
        from . import runtime_settings
        self.rs = runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.owner = User.objects.create_user('owner', 'owner@example.com', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Bücher')
        self.item = Item.objects.create(collection=self.col, values={'name': 'Dune'})

    def _overdue_loan(self, **kwargs):
        from datetime import timedelta
        from django.utils import timezone
        defaults = {'item': self.item, 'borrower': 'Anna',
                    'lent_at': timezone.localdate() - timedelta(days=5),
                    'due_at': timezone.localdate() - timedelta(days=1)}
        defaults.update(kwargs)
        return Loan.objects.create(**defaults)

    def _run(self):
        from django.core.management import call_command
        call_command('send_loan_reminders', stdout=__import__('io').StringIO())

    def test_disabled_sends_nothing(self):
        from django.core import mail
        self._overdue_loan()
        self._run()
        self.assertEqual(len(mail.outbox), 0)

    def test_overdue_loan_mails_owner_and_stamps(self):
        from django.core import mail
        self.rs.set_setting('loan_reminders_enabled', True)
        loan = self._overdue_loan()
        self._run()
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ['owner@example.com'])
        self.assertIn('Dune', mail.outbox[0].body)
        self.assertIn('Anna', mail.outbox[0].body)
        loan.refresh_from_db()
        self.assertIsNotNone(loan.reminder_sent_at)

    def test_loan_without_due_date_uses_overdue_days(self):
        from datetime import timedelta
        from django.core import mail
        from django.utils import timezone
        self.rs.set_setting('loan_reminders_enabled', True)
        self._overdue_loan(due_at=None,
                           lent_at=timezone.localdate() - timedelta(days=31))
        self._run()
        self.assertEqual(len(mail.outbox), 1)

    def test_not_overdue_not_mailed(self):
        from datetime import timedelta
        from django.core import mail
        from django.utils import timezone
        self.rs.set_setting('loan_reminders_enabled', True)
        self._overdue_loan(due_at=timezone.localdate() + timedelta(days=3))
        self._run()
        self.assertEqual(len(mail.outbox), 0)

    def test_no_resend_within_interval_but_after(self):
        from datetime import timedelta
        from django.core import mail
        from django.utils import timezone
        self.rs.set_setting('loan_reminders_enabled', True)
        loan = self._overdue_loan()
        self._run()
        self._run()  # immediately again: still 1 mail
        self.assertEqual(len(mail.outbox), 1)
        Loan.objects.filter(pk=loan.pk).update(
            reminder_sent_at=timezone.now() - timedelta(days=8))
        self._run()
        self.assertEqual(len(mail.outbox), 2)

    def test_returned_and_trashed_loans_skipped(self):
        from django.core import mail
        from django.utils import timezone
        self.rs.set_setting('loan_reminders_enabled', True)
        self._overdue_loan(returned_at=timezone.localdate())
        other_item = Item.objects.create(collection=self.col, values={'name': 'Weg'})
        self._overdue_loan(item=other_item)
        other_item.soft_delete()
        self._run()
        self.assertEqual(len(mail.outbox), 0)

    def test_one_digest_per_owner(self):
        from django.core import mail
        self.rs.set_setting('loan_reminders_enabled', True)
        second = Item.objects.create(collection=self.col, values={'name': 'Hyperion'})
        self._overdue_loan()
        self._overdue_loan(item=second, borrower='Ben')
        self._run()
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn('Dune', mail.outbox[0].body)
        self.assertIn('Hyperion', mail.outbox[0].body)


class ApiTests(TestCase):
    """The token-authenticated JSON API (gated by api_enabled)."""

    def setUp(self):
        from django.core.cache import cache
        from accounts.models import ApiToken
        from . import runtime_settings
        self.rs = runtime_settings
        cache.delete(runtime_settings._CACHE_KEY)
        self.addCleanup(cache.delete, runtime_settings._CACHE_KEY)
        self.rs.set_setting('api_enabled', True)

        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Bücher')
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0, required=True)
        add_field(self.col, 'ort', 'Ort', FieldType.TEXT, order=1)
        CollectionShare.objects.create(collection=self.col, user=self.viewer, permission='view')
        self.token, self.token_key = ApiToken.create_for_user(self.owner, 'Test')
        self.viewer_token, self.viewer_key = ApiToken.create_for_user(self.viewer, 'Viewer')
        self.auth = {'headers': {'Authorization': f'Bearer {self.token_key}'}}
        self.items_url = reverse('api_items', args=[self.col.pk])

    def test_disabled_api_rejects_valid_token(self):
        self.rs.set_setting('api_enabled', False)
        resp = self.client.get(reverse('api_collections'), **self.auth)
        self.assertEqual(resp.status_code, 403)

    def test_missing_and_invalid_token(self):
        self.assertEqual(self.client.get(reverse('api_collections')).status_code, 401)
        resp = self.client.get(reverse('api_collections'),
                               headers={'Authorization': 'Bearer falsch'})
        self.assertEqual(resp.status_code, 401)

    def test_inactive_user_token_rejected(self):
        self.owner.is_active = False
        self.owner.save(update_fields=['is_active'])
        self.assertEqual(self.client.get(reverse('api_collections'), **self.auth).status_code, 401)

    def test_list_collections_with_permission(self):
        resp = self.client.get(reverse('api_collections'), **self.auth)
        self.assertEqual(resp.status_code, 200)
        data = resp.json()['results']
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['name'], 'Bücher')
        self.assertEqual(data[0]['permission'], 'owner')

    def test_foreign_collection_forbidden(self):
        stranger = User.objects.create_user('fremd', 'f@e.de', 'pw')
        foreign = Collection.objects.create(owner=stranger, name='Privat')
        resp = self.client.get(reverse('api_collection_detail', args=[foreign.pk]), **self.auth)
        self.assertEqual(resp.status_code, 403)

    def test_collection_schema(self):
        ItemType.objects.create(collection=self.col, name='Roman')
        resp = self.client.get(reverse('api_collection_detail', args=[self.col.pk]), **self.auth)
        data = resp.json()
        self.assertEqual([f['key'] for f in data['fields']], ['name', 'ort'])
        self.assertTrue(data['fields'][0]['required'])
        self.assertEqual(data['item_types'][0]['name'], 'Roman')

    def test_items_list_filter_and_pagination(self):
        for name in ('Dune', 'Hyperion'):
            Item.objects.create(collection=self.col, values={'name': name})
        resp = self.client.get(self.items_url, {'q': 'dune'}, **self.auth)
        data = resp.json()
        self.assertEqual(data['count'], 1)
        self.assertEqual(data['page'], 1)
        self.assertEqual(data['results'][0]['values']['name'], 'Dune')

    def test_create_item(self):
        art = ItemType.objects.create(collection=self.col, name='Roman')
        resp = self.client.post(
            self.items_url, data='{"values": {"name": "Dune"}, "item_type": %d}' % art.pk,
            content_type='application/json', **self.auth)
        self.assertEqual(resp.status_code, 201)
        item = self.col.items.get()
        self.assertEqual(item.values['name'], 'Dune')
        self.assertEqual(item.item_type, art)
        self.assertEqual(item.created_by, self.owner)

    def test_create_validates_required_fields(self):
        resp = self.client.post(self.items_url, data='{"values": {"ort": "Regal"}}',
                                content_type='application/json', **self.auth)
        self.assertEqual(resp.status_code, 400)
        self.assertIn('name', resp.json()['fields'])
        self.assertEqual(self.col.items.count(), 0)

    def test_invalid_json_rejected(self):
        resp = self.client.post(self.items_url, data='kein json',
                                content_type='application/json', **self.auth)
        self.assertEqual(resp.status_code, 400)

    def test_patch_merges_put_replaces(self):
        item = Item.objects.create(collection=self.col,
                                   values={'name': 'Dune', 'ort': 'Regal A'})
        url = reverse('api_item', args=[self.col.pk, item.pk])
        resp = self.client.patch(url, data='{"values": {"ort": "Regal B"}}',
                                 content_type='application/json', **self.auth)
        self.assertEqual(resp.status_code, 200)
        item.refresh_from_db()
        self.assertEqual(item.values['name'], 'Dune')       # kept (merge)
        self.assertEqual(item.values['ort'], 'Regal B')

        resp = self.client.put(url, data='{"values": {"name": "Duna"}}',
                               content_type='application/json', **self.auth)
        self.assertEqual(resp.status_code, 200)
        item.refresh_from_db()
        self.assertEqual(item.values['name'], 'Duna')
        self.assertFalse(item.values.get('ort'))            # replaced (PUT)

    def test_delete_moves_to_trash(self):
        item = Item.objects.create(collection=self.col, values={'name': 'Weg'})
        url = reverse('api_item', args=[self.col.pk, item.pk])
        resp = self.client.delete(url, **self.auth)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(self.col.items.count(), 0)
        self.assertTrue(Item.all_objects.filter(pk=item.pk,
                                                deleted_at__isnull=False).exists())

    def test_viewer_can_read_but_not_write(self):
        Item.objects.create(collection=self.col, values={'name': 'Dune'})
        viewer_auth = {'headers': {'Authorization': f'Bearer {self.viewer_key}'}}
        self.assertEqual(self.client.get(self.items_url, **viewer_auth).status_code, 200)
        resp = self.client.post(self.items_url, data='{"values": {"name": "Nein"}}',
                                content_type='application/json', **viewer_auth)
        self.assertEqual(resp.status_code, 403)

    def test_token_use_is_stamped(self):
        self.assertIsNone(self.token.last_used_at)
        self.client.get(reverse('api_collections'), **self.auth)
        self.token.refresh_from_db()
        self.assertIsNotNone(self.token.last_used_at)


class SecurityHeaderTests(TestCase):
    """Crawler protection and browser hardening headers on every response."""

    def test_robots_txt_disallows_everything(self):
        resp = self.client.get('/robots.txt')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'text/plain')
        self.assertContains(resp, 'Disallow: /')

    def test_x_robots_tag_and_permissions_policy_on_all_pages(self):
        resp = self.client.get(reverse('login'))
        self.assertIn('noindex', resp.headers.get('X-Robots-Tag', ''))
        self.assertIn('camera=(self)', resp.headers.get('Permissions-Policy', ''))

    def test_meta_robots_in_html(self):
        resp = self.client.get(reverse('login'))
        self.assertContains(resp, '<meta name="robots" content="noindex, nofollow">')

    def test_content_security_policy_enforced_with_nonce(self):
        resp = self.client.get(reverse('login'))
        policy = resp.headers.get('Content-Security-Policy', '')
        self.assertIn("default-src 'self'", policy)
        self.assertIn("frame-ancestors 'none'", policy)
        self.assertIn('nonce-', policy)  # script nonce present …
        import re
        nonce = re.search(r"'nonce-([^']+)'", policy).group(1)
        self.assertContains(resp, 'nonce="%s"' % nonce)  # … and used inline

    def test_no_third_party_asset_urls_in_pages(self):
        # GDPR: no request to CDNs & Co. may be triggered by our pages.
        resp = self.client.get(reverse('login'))
        self.assertNotContains(resp, 'cdn.jsdelivr.net')

    def test_clickjacking_and_sniffing_headers(self):
        resp = self.client.get(reverse('login'))
        self.assertEqual(resp.headers.get('X-Frame-Options'), 'DENY')
        self.assertEqual(resp.headers.get('X-Content-Type-Options'), 'nosniff')
        self.assertEqual(resp.headers.get('Referrer-Policy'), 'same-origin')


class LookupThrottleTests(TestCase):
    """External-database lookups are capped per user (outbound-request abuse)."""

    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.coll = Collection.objects.create(owner=self.owner, name='Bücher')
        add_field(self.coll, 'isbn', 'ISBN', FieldType.ISBN, order=0,
                  config={'lookup_attribute': 'isbn'})
        self.client.force_login(self.owner)

    def test_lookup_returns_429_after_limit(self):
        url = reverse('item_lookup', args=[self.coll.pk])
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_EMPTY_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json', return_value={}):
            for _ in range(30):
                self.assertEqual(self.client.get(url, {'q': '123'}).status_code, 200)
            resp = self.client.get(url, {'q': '123'})
        self.assertEqual(resp.status_code, 429)
        self.assertFalse(resp.json()['ok'])

    def test_search_shares_the_same_user_budget(self):
        lookup_url = reverse('item_lookup', args=[self.coll.pk])
        search_url = reverse('item_search', args=[self.coll.pk])
        with mock.patch.object(lookup_providers, '_http_get_text', return_value=DNB_EMPTY_PAYLOAD), \
                mock.patch.object(lookup_providers, '_http_get_json', return_value={}):
            for _ in range(30):
                self.client.get(lookup_url, {'q': '123'})
            resp = self.client.get(search_url, {'q': 'hawking'})
        self.assertEqual(resp.status_code, 429)


# --- Media providers beyond books (music, movies, games, board games, EAN) -----

MUSICBRAINZ_PAYLOAD = {'releases': [{
    'id': 'mbid-123',
    'title': 'Nevermind',
    'artist-credit': [{'name': 'Nirvana', 'joinphrase': ''}],
    'date': '1991-09-24',
    'barcode': '0720642442524',
    'label-info': [{'label': {'name': 'DGC'}}],
    'media': [{'format': 'CD', 'track-count': 12}],
}]}

TMDB_PAYLOAD = {'results': [
    {'media_type': 'movie', 'title': 'Inception', 'release_date': '2010-07-15',
     'overview': 'Ein Dieb stiehlt Geheimnisse aus Träumen.',
     'poster_path': '/abc.jpg', 'genre_ids': [28, 878], 'original_language': 'en'},
    {'media_type': 'person', 'name': 'Christopher Nolan'},
]}

RAWG_PAYLOAD = {'results': [
    {'name': 'The Legend of Zelda: Breath of the Wild', 'released': '2017-03-03',
     'background_image': 'https://media.rawg.io/media/zelda.jpg',
     'platforms': [{'platform': {'name': 'Nintendo Switch'}}, {'platform': {'name': 'Wii U'}}],
     'genres': [{'name': 'Adventure'}, {'name': 'Action'}]},
]}

WIKIDATA_SEARCH_PAYLOAD = {'search': [
    {'id': 'Q1903', 'label': 'Catania', 'description': 'Stadt in Sizilien'},
    {'id': 'Q17271', 'label': 'Die Siedler von Catan', 'description': 'Brettspiel'},
]}

WIKIDATA_ENTITIES_PAYLOAD = {'entities': {
    'Q1903': {  # the Sicilian city — must be filtered out via P31
        'labels': {'de': {'value': 'Catania'}},
        'claims': {'P31': [{'mainsnak': {'datavalue': {'value': {'id': 'Q515'}}}}]},
    },
    'Q17271': {
        'labels': {'de': {'value': 'Die Siedler von Catan'}},
        'descriptions': {'de': {'value': 'Brettspiel'}},
        'claims': {
            'P31': [{'mainsnak': {'datavalue': {'value': {'id': 'Q131436'}}}}],
            'P577': [{'mainsnak': {'datavalue': {'value': {'time': '+1995-00-00T00:00:00Z'}}}}],
            'P1872': [{'mainsnak': {'datavalue': {'value': {'amount': '+3'}}}}],
            'P1873': [{'mainsnak': {'datavalue': {'value': {'amount': '+4'}}}}],
            'P18': [{'mainsnak': {'datavalue': {'value': 'Catan Brettspiel.jpg'}}}],
        },
    },
}}

UPC_PAYLOAD = {'code': 'OK', 'items': [{
    'ean': '0885909950805', 'title': 'Apple iPhone', 'brand': 'Apple',
    'category': 'Electronics > Communications', 'description': 'x' * 700,
    'images': ['https://some-shop.example/iphone.jpg'],
}]}


class MediaProviderTests(TestCase):
    def _set_key(self, key, value):
        from .runtime_settings import _CACHE_KEY, set_setting
        set_setting(key, value)
        self.addCleanup(cache.delete, _CACHE_KEY)

    def test_musicbrainz_fetch_parses_release_and_cover(self):
        with mock.patch.object(lookup_providers, '_http_get_json',
                               return_value=MUSICBRAINZ_PAYLOAD) as m:
            data = lookup_providers.get_provider('musicbrainz').fetch('0-720642-442524')
        self.assertIn('barcode%3A0720642442524', m.call_args[0][0])
        self.assertEqual(data['title'], 'Nevermind')
        self.assertEqual(data['artist'], 'Nirvana')
        self.assertEqual(data['authors'], 'Nirvana')
        self.assertEqual(data['publisher'], 'DGC')
        self.assertEqual(data['year'], '1991')
        self.assertEqual(data['format'], 'CD')
        self.assertEqual(data['ean'], '0720642442524')
        self.assertEqual(data['cover_url'],
                         'https://coverartarchive.org/release/mbid-123/front-250')

    def test_musicbrainz_search_escapes_lucene_specials(self):
        with mock.patch.object(lookup_providers, '_http_get_json',
                               return_value=MUSICBRAINZ_PAYLOAD) as m:
            results = lookup_providers.get_provider('musicbrainz').search('nirvana: nevermind!')
        self.assertNotIn('%3A', m.call_args[0][0].split('query=')[1].split('&')[0])
        self.assertEqual(results[0]['title'], 'Nevermind')

    def test_tmdb_requires_api_key(self):
        provider = lookup_providers.get_provider('tmdb')
        self.assertFalse(provider.is_available())
        self.assertEqual(provider.search('inception'), [])

    def test_tmdb_search_with_key_parses_movies_only(self):
        self._set_key('tmdb_api_key', 'k123')
        provider = lookup_providers.get_provider('tmdb')
        self.assertTrue(provider.is_available())
        with mock.patch.object(lookup_providers, '_http_get_json',
                               return_value=TMDB_PAYLOAD) as m:
            results = provider.search('inception')
        self.assertIn('api_key=k123', m.call_args[0][0])
        self.assertEqual(len(results), 1)  # the person entry is dropped
        self.assertEqual(results[0]['title'], 'Inception')
        self.assertEqual(results[0]['year'], '2010')
        self.assertEqual(results[0]['categories'], 'Action, Science Fiction')
        self.assertEqual(results[0]['language'], 'Englisch')
        self.assertEqual(results[0]['cover_url'], 'https://image.tmdb.org/t/p/w342/abc.jpg')

    def test_rawg_search_with_key_parses_games(self):
        self._set_key('rawg_api_key', 'r123')
        with mock.patch.object(lookup_providers, '_http_get_json', return_value=RAWG_PAYLOAD):
            results = lookup_providers.get_provider('rawg').search('zelda')
        self.assertEqual(results[0]['title'], 'The Legend of Zelda: Breath of the Wild')
        self.assertEqual(results[0]['platform'], 'Nintendo Switch, Wii U')
        self.assertEqual(results[0]['categories'], 'Adventure, Action')
        self.assertEqual(results[0]['cover_url'], 'https://media.rawg.io/media/zelda.jpg')

    def test_wikidata_search_filters_to_games_and_parses_claims(self):
        with mock.patch.object(lookup_providers, '_http_get_json',
                               side_effect=[WIKIDATA_SEARCH_PAYLOAD,
                                            WIKIDATA_ENTITIES_PAYLOAD]) as m:
            results = lookup_providers.get_provider('wikidata').search('catan')
        self.assertEqual(m.call_count, 2)
        self.assertIn('Q1903%7CQ17271', m.call_args[0][0])  # one batched claims call
        self.assertEqual(len(results), 1)  # the city is filtered out via P31
        game = results[0]
        self.assertEqual(game['title'], 'Die Siedler von Catan')
        self.assertEqual(game['year'], '1995')
        self.assertEqual(game['players'], '3–4')
        self.assertEqual(game['description'], 'Brettspiel')
        self.assertEqual(game['cover_url'],
                         'https://commons.wikimedia.org/wiki/Special:FilePath/'
                         'Catan%20Brettspiel.jpg?width=400')

    def test_upcitemdb_lookup_never_emits_untrusted_cover(self):
        with mock.patch.object(lookup_providers, '_http_get_json', return_value=UPC_PAYLOAD):
            data = lookup_providers.get_provider('upcitemdb').fetch('0885909950805')
        self.assertEqual(data['title'], 'Apple iPhone')
        self.assertEqual(data['brand'], 'Apple')
        self.assertEqual(len(data['description']), 600)
        self.assertNotIn('cover_url', data)  # merchant image hosts are untrusted

    def test_chain_for_music_and_availability_gating(self):
        keys = [p.key for p in lookup_providers.chain_for('music')]
        self.assertEqual(keys, ['musicbrainz', 'upcitemdb'])
        keys_all = [p.key for p in lookup_providers.chain_for('')]
        self.assertNotIn('tmdb', keys_all)  # no API key configured
        self._set_key('tmdb_api_key', 'k123')
        keys_all = [p.key for p in lookup_providers.chain_for('')]
        self.assertIn('tmdb', keys_all)

    def test_provider_for_uses_collection_media_kind(self):
        owner = User.objects.create_user('kindowner', 'k@e.de', 'pw')
        coll = Collection.objects.create(owner=owner, name='Platten', lookup_provider='music')
        provider = lookup_providers.provider_for(coll)
        self.assertIn('MusicBrainz', provider.label)
        self.assertNotIn('DNB', provider.label)
        self.assertEqual(provider.query_attribute, 'ean')

    def test_ean_fetch_prefers_product_sources_over_book_sources(self):
        calls = []

        def fake_json(url):
            calls.append(url)
            if 'musicbrainz' in url:
                return MUSICBRAINZ_PAYLOAD
            return {}

        with mock.patch.object(lookup_providers, '_http_get_json', side_effect=fake_json), \
                mock.patch.object(lookup_providers, '_http_get_text', return_value=None):
            data = lookup_providers.auto_provider().fetch('0720642442524')
        self.assertEqual(data['title'], 'Nevermind')
        self.assertIn('musicbrainz', calls[0])  # non-ISBN code: product/music sources first

    def test_cover_hosts_cover_new_providers(self):
        for url in ('https://coverartarchive.org/release/x/front-250',
                    'https://ia800505.us.archive.org/x.jpg',
                    'https://image.tmdb.org/t/p/w342/abc.jpg',
                    'https://media.rawg.io/media/zelda.jpg',
                    'https://commons.wikimedia.org/wiki/Special:FilePath/x.jpg',
                    'https://upload.wikimedia.org/wikipedia/commons/x.jpg'):
            self.assertTrue(lookup_providers.cover_url_allowed(url), url)
        self.assertFalse(lookup_providers.cover_url_allowed('https://some-shop.example/i.jpg'))
        self.assertFalse(lookup_providers.cover_url_allowed('https://evilarchive.org/x'))

    def test_lookup_view_requires_login(self):
        owner = User.objects.create_user('lockowner', 'l@e.de', 'pw')
        coll = Collection.objects.create(owner=owner, name='L')
        resp = self.client.get(reverse('item_lookup', args=[coll.pk]), {'q': '1'})
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/accounts/login/', resp['Location'])


class MediaKindTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('mkowner', 'mk@e.de', 'pw')
        self.client.force_login(self.owner)

    def test_presets_store_their_media_kind(self):
        from .services import PRESETS
        for preset, expected in (('books', 'books'), ('movies', 'movies'),
                                 ('music', 'music'), ('games', 'games'),
                                 ('boardgames', 'boardgames'), ('generic', '')):
            name = 'Kind %s' % preset
            self.client.post(reverse('collection_create'),
                             {'name': name, 'description': '', 'preset': preset})
            coll = Collection.objects.get(name=name)
            self.assertEqual(coll.lookup_provider, expected, preset)
            self.assertIn(preset, PRESETS)

    def test_movie_preset_maps_lookup_attributes(self):
        self.client.post(reverse('collection_create'),
                         {'name': 'Filme', 'description': '', 'preset': 'movies'})
        coll = Collection.objects.get(name='Filme')
        self.assertEqual(coll.fields.get(key='regie').config.get('lookup_attribute'), 'director')
        self.assertEqual(coll.fields.get(key='ean').config.get('lookup_attribute'), 'ean')
        self.assertEqual(coll.fields.get(key='bild').config.get('lookup_attribute'), 'cover_url')

    def test_edit_form_changes_media_kind(self):
        coll = Collection.objects.create(owner=self.owner, name='Regal')
        resp = self.client.post(reverse('collection_edit', args=[coll.pk]),
                                {'name': 'Regal', 'description': '', 'lookup_provider': 'music'})
        self.assertEqual(resp.status_code, 302)
        coll.refresh_from_db()
        self.assertEqual(coll.lookup_provider, 'music')

    def test_copy_structure_copies_media_kind(self):
        from .services import copy_structure
        source = Collection.objects.create(owner=self.owner, name='Q', lookup_provider='games')
        target = Collection.objects.create(owner=self.owner, name='Z')
        copy_structure(source, target)
        target.refresh_from_db()
        self.assertEqual(target.lookup_provider, 'games')


# --- Price comparison / multi-platform search ----------------------------------

class PriceSearchBuildTests(TestCase):
    def _links(self, **kwargs):
        from . import price_search
        return {entry['platform'].key: entry
                for entry in price_search.build_links(price_search.PriceQuery(**kwargs))}

    def test_book_isbn_builds_precise_links(self):
        links = self._links(q='Der Hobbit Tolkien', code='978-3-608-93981-1', kind='books')
        self.assertIn('eurobuch', links)
        self.assertIn('/buch/isbn/9783608939811.html', links['eurobuch']['url'])
        self.assertIn('isbn=9783608939811', links['booklooker']['url'])
        self.assertIn('_nkw=9783608939811', links['ebay']['url'])
        self.assertIn('i=stripbooks', links['amazon']['url'])
        self.assertTrue(links['idealo']['precise'])
        self.assertNotIn('discogs', links)  # music only

    def test_condition_and_price_filters(self):
        from decimal import Decimal
        links = self._links(q='Der Hobbit', kind='books', condition='new',
                            min_price=Decimal('5'), max_price=Decimal('20'), sort='price')
        self.assertNotIn('medimops', links)   # used-only platform filtered out
        self.assertNotIn('booklooker', links)
        self.assertIn('thalia', links)
        ebay = links['ebay']['url']
        self.assertIn('LH_ItemCondition=1000', ebay)
        self.assertIn('_udlo=5', ebay)
        self.assertIn('_udhi=20', ebay)
        self.assertIn('_sop=15', ebay)
        used = self._links(q='Der Hobbit', kind='books', condition='used')
        self.assertIn('medimops', used)
        self.assertNotIn('thalia', used)

    def test_music_ean_uses_discogs_barcode_search(self):
        links = self._links(code='0720642442524', kind='music')
        self.assertIn('barcode=0720642442524', links['discogs']['url'])
        self.assertIn('medimops', links)
        self.assertNotIn('booklooker', links)

    def test_no_query_returns_nothing(self):
        from . import price_search
        self.assertEqual(price_search.build_links(price_search.PriceQuery()), [])

    def test_eurobuch_needs_isbn(self):
        links = self._links(q='irgendwas', kind='books')
        self.assertNotIn('eurobuch', links)


class PriceSearchViewTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('priceowner', 'p@e.de', 'pw')
        self.other = User.objects.create_user('priceother', 'po@e.de', 'pw')
        self.client.force_login(self.owner)
        self.coll = Collection.objects.create(owner=self.owner, name='Bücher',
                                              lookup_provider='books')
        add_field(self.coll, 'titel', 'Titel', FieldType.TEXT, 0,
                  config={'lookup_attribute': 'title'})
        add_field(self.coll, 'autor', 'Autor', FieldType.TEXT, 1,
                  config={'lookup_attribute': 'authors'})
        add_field(self.coll, 'isbn', 'ISBN', FieldType.ISBN, 2,
                  config={'lookup_attribute': 'isbn'})
        self.item = Item.objects.create(collection=self.coll, values={
            'titel': 'Der Hobbit', 'autor': 'J.R.R. Tolkien', 'isbn': '978-3-608-93981-1'})

    def test_page_requires_login(self):
        self.client.logout()
        resp = self.client.get(reverse('price_search'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/accounts/login/', resp['Location'])

    def test_page_renders_grouped_platform_links(self):
        resp = self.client.get(reverse('price_search'),
                               {'q': 'Der Hobbit', 'code': '9783608939811', 'kind': 'books'})
        self.assertContains(resp, 'eurobuch.com')
        self.assertContains(resp, 'rel="noopener noreferrer nofollow"')
        self.assertContains(resp, 'Preisvergleich &amp; Metasuche')
        self.assertContains(resp, 'Gebraucht kaufen')

    def test_item_button_prefills_from_item(self):
        resp = self.client.get(reverse('item_price_search', args=[self.coll.pk, self.item.pk]))
        self.assertEqual(resp.status_code, 302)
        location = resp['Location']
        self.assertIn('code=978-3-608-93981-1', location.replace('%2D', '-'))
        self.assertIn('kind=books', location)
        follow = self.client.get(location)
        self.assertContains(follow, 'eurobuch.com')
        self.assertContains(follow, 'Der+Hobbit+J.R.R.+Tolkien')

    def test_item_price_search_respects_row_level_permission(self):
        self.client.force_login(self.other)
        resp = self.client.get(reverse('item_price_search', args=[self.coll.pk, self.item.pk]))
        self.assertEqual(resp.status_code, 403)

    def test_item_detail_links_to_price_search(self):
        resp = self.client.get(reverse('item_detail', args=[self.coll.pk, self.item.pk]))
        self.assertContains(resp, reverse('item_price_search', args=[self.coll.pk, self.item.pk]))


# --- GDPR: legal pages, data export, account deletion ---------------------------

class LegalPagesTests(TestCase):
    def test_privacy_and_imprint_are_public(self):
        for name in ('privacy', 'imprint'):
            resp = self.client.get(reverse(name))
            self.assertEqual(resp.status_code, 200, name)
        self.assertContains(self.client.get(reverse('privacy')), 'Datenschutzerklärung')
        self.assertContains(self.client.get(reverse('imprint')), 'Impressum')

    def test_operator_details_come_from_settings(self):
        from .runtime_settings import _CACHE_KEY, set_setting
        set_setting('legal_operator', 'Max Mustermann')
        set_setting('legal_address', 'Musterweg 1\n12345 Musterstadt')
        set_setting('legal_email', 'max@example.org')
        self.addCleanup(cache.delete, _CACHE_KEY)
        resp = self.client.get(reverse('imprint'))
        self.assertContains(resp, 'Max Mustermann')
        self.assertContains(resp, 'Musterweg 1')
        self.assertContains(resp, 'max@example.org')

    def test_legal_pages_reachable_in_maintenance_mode(self):
        from .runtime_settings import _CACHE_KEY, set_setting
        set_setting('maintenance_mode', True)
        self.addCleanup(cache.delete, _CACHE_KEY)
        self.assertEqual(self.client.get(reverse('privacy')).status_code, 200)
        self.assertEqual(self.client.get(reverse('imprint')).status_code, 200)

    def test_footer_links_present(self):
        user = User.objects.create_user('legaluser', 'lg@e.de', 'pw')
        self.client.force_login(user)
        resp = self.client.get(reverse('dashboard'))
        self.assertContains(resp, reverse('privacy'))
        self.assertContains(resp, reverse('imprint'))


@override_settings(MEDIA_ROOT=MEDIA)
class DataExportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('exporter', 'ex@e.de', 'pw',
                                             display_name='Exporteur')
        self.client.force_login(self.user)
        self.coll = Collection.objects.create(owner=self.user, name='Meine Bücher')
        add_field(self.coll, 'titel', 'Titel', FieldType.TEXT)
        self.item = Item.objects.create(collection=self.coll, values={'titel': 'Der Hobbit'})
        Loan.objects.create(item=self.item, borrower='Anna')

    def test_export_contains_account_and_collection_data(self):
        resp = self.client.get(reverse('data_export'))
        self.assertEqual(resp.status_code, 200)
        self.assertIn('attachment', resp['Content-Disposition'])
        data = json.loads(resp.content)
        self.assertEqual(data['account']['username'], 'exporter')
        self.assertEqual(data['account']['display_name'], 'Exporteur')
        collection = data['collections'][0]
        self.assertEqual(collection['name'], 'Meine Bücher')
        self.assertEqual(collection['items'][0]['values']['titel'], 'Der Hobbit')
        self.assertEqual(collection['items'][0]['loans'][0]['borrower'], 'Anna')

    def test_export_requires_login(self):
        self.client.logout()
        resp = self.client.get(reverse('data_export'))
        self.assertEqual(resp.status_code, 302)


@override_settings(MEDIA_ROOT=MEDIA)
class AccountDeleteTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('deleteme', 'del@e.de', 'pw')
        self.client.force_login(self.user)
        self.coll = Collection.objects.create(owner=self.user, name='Weg damit')
        add_field(self.coll, 'bild', 'Bild', FieldType.IMAGE)

    def _upload_item(self):
        resp = self.client.post(
            reverse('item_create', args=[self.coll.pk]),
            {'bild': SimpleUploadedFile('foto.png', make_png(), 'image/png')})
        self.assertEqual(resp.status_code, 302)
        return self.coll.items.first().assets.first()

    def test_confirmation_page_shows_counts(self):
        Item.objects.create(collection=self.coll, values={})
        resp = self.client.get(reverse('account_delete'))
        self.assertContains(resp, 'Konto endgültig löschen')
        self.assertContains(resp, '1 Sammlung')
        self.assertContains(resp, '1 Gegenstand')

    def test_wrong_password_keeps_account(self):
        self.client.post(reverse('account_delete'), {'password': 'falsch'})
        self.assertTrue(User.objects.filter(username='deleteme').exists())

    def test_delete_removes_account_collections_and_files(self):
        import os
        asset = self._upload_item()
        path = asset.file.path
        self.assertTrue(os.path.exists(path))
        resp = self.client.post(reverse('account_delete'), {'password': 'pw'})
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse('login'), resp['Location'])
        self.assertFalse(User.objects.filter(username='deleteme').exists())
        self.assertFalse(Collection.objects.filter(name='Weg damit').exists())
        self.assertFalse(os.path.exists(path))

    def test_sole_superuser_cannot_self_delete(self):
        admin = User.objects.create_superuser('lonelyadmin', 'a@e.de', 'pw')
        self.client.force_login(admin)
        resp = self.client.get(reverse('account_delete'))
        self.assertContains(resp, 'einzige aktive Administrator')
        self.client.post(reverse('account_delete'), {'password': 'pw'})
        self.assertTrue(User.objects.filter(username='lonelyadmin').exists())

    def test_items_created_in_foreign_collections_are_anonymised(self):
        other = User.objects.create_user('keeper', 'k@e.de', 'pw')
        their = Collection.objects.create(owner=other, name='Bleibt')
        foreign_item = Item.objects.create(collection=their, values={}, created_by=self.user)
        self.client.post(reverse('account_delete'), {'password': 'pw'})
        foreign_item.refresh_from_db()
        self.assertIsNone(foreign_item.created_by)


class UrlValueHardeningTests(TestCase):
    """URL-typed values: only real web URLs become links (import + rendering)."""

    def setUp(self):
        self.owner = User.objects.create_user('urlowner', 'u@e.de', 'pw')
        self.client.force_login(self.owner)
        self.coll = Collection.objects.create(owner=self.owner, name='Links')
        self.fd = add_field(self.coll, 'link', 'Link', FieldType.URL)

    def test_render_cell_refuses_non_web_schemes(self):
        from .rendering import render_cell
        self.assertEqual(render_cell(self.fd, 'https://example.org/x').kind, 'url')
        for bad in ('javascript:alert(1)', 'data:text/html,x', 'ftp://example.org'):
            cell = render_cell(self.fd, bad)
            self.assertEqual(cell.kind, 'text', bad)
            self.assertEqual(cell.url, '')

    def test_import_validates_url_scheme(self):
        result = imports.import_table(self.coll, [
            ['Link'],
            ['https://example.org/a'],
            ['www.example.org/b'],           # bare domain: scheme added
            ['javascript:alert(1)'],         # refused with a row warning
        ], user=self.owner)
        values = sorted(item.values.get('link', '') for item in self.coll.items.all())
        self.assertIn('https://example.org/a', values)
        self.assertIn('https://www.example.org/b', values)
        self.assertFalse(any('javascript' in v for v in values))
        self.assertEqual(len(result['warnings']), 1)

    def test_item_detail_links_carry_noopener(self):
        item = Item.objects.create(collection=self.coll,
                                   values={'link': 'https://example.org/x'})
        resp = self.client.get(reverse('item_detail', args=[self.coll.pk, item.pk]))
        self.assertContains(resp, 'rel="noopener noreferrer">https://example.org/x')


class DataExportThrottleTests(TestCase):
    def test_export_is_rate_limited(self):
        user = User.objects.create_user('throttled', 't@e.de', 'pw')
        self.client.force_login(user)
        for _ in range(10):
            self.assertEqual(self.client.get(reverse('data_export')).status_code, 200)
        self.assertEqual(self.client.get(reverse('data_export')).status_code, 429)


class ItemBrowseNavigationTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('navowner', 'n@e.de', 'pw')
        self.client.force_login(self.owner)
        self.coll = Collection.objects.create(owner=self.owner, name='Nav')
        add_field(self.coll, 'name', 'Name', FieldType.TEXT)
        self.a = Item.objects.create(collection=self.coll, values={'name': 'Ältester'})
        self.b = Item.objects.create(collection=self.coll, values={'name': 'Mitte'})
        self.c = Item.objects.create(collection=self.coll, values={'name': 'Neuester'})

    def test_middle_item_links_to_both_neighbours(self):
        resp = self.client.get(reverse('item_detail', args=[self.coll.pk, self.b.pk]))
        self.assertContains(resp, reverse('item_detail', args=[self.coll.pk, self.c.pk]))
        self.assertContains(resp, reverse('item_detail', args=[self.coll.pk, self.a.pk]))

    def test_newest_item_has_no_newer_link(self):
        resp = self.client.get(reverse('item_detail', args=[self.coll.pk, self.c.pk]))
        # Only the next-older neighbour is linked; the oldest item is not.
        self.assertNotContains(resp, reverse('item_detail', args=[self.coll.pk, self.a.pk]))
        self.assertContains(resp, reverse('item_detail', args=[self.coll.pk, self.b.pk]))

    def test_lend_due_date_has_min_today(self):
        resp = self.client.get(reverse('item_detail', args=[self.coll.pk, self.b.pk]))
        from django.utils import timezone
        self.assertContains(resp, 'min="%s"' % timezone.localdate().isoformat())


class PerPageOverrideTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('ppowner', 'pp@e.de', 'pw')
        self.client.force_login(self.owner)
        self.coll = Collection.objects.create(owner=self.owner, name='Viele')
        add_field(self.coll, 'name', 'Name', FieldType.TEXT)
        Item.objects.bulk_create([
            Item(collection=self.coll, values={'name': f'Ding {i}'}) for i in range(30)
        ])

    def test_per_page_query_param_overrides_setting(self):
        resp = self.client.get(reverse('collection_detail', args=[self.coll.pk]),
                               {'per_page': 25})
        self.assertEqual(len(resp.context['rows']), 25)
        self.assertContains(resp, 'Pro Seite:')
        self.assertContains(resp, '<strong>25</strong>')

    def test_invalid_per_page_falls_back_to_setting(self):
        resp = self.client.get(reverse('collection_detail', args=[self.coll.pk]),
                               {'per_page': 'kaputt'})
        self.assertEqual(len(resp.context['rows']), 30)  # default 50 per page

    def test_per_page_is_clamped(self):
        resp = self.client.get(reverse('collection_detail', args=[self.coll.pk]),
                               {'per_page': 1})
        self.assertEqual(len(resp.context['rows']), 5)  # lower clamp


class SavedViewTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('svowner', 'sv@e.de', 'pw')
        self.viewer = User.objects.create_user('svviewer', 'svv@e.de', 'pw')
        self.client.force_login(self.owner)
        self.coll = Collection.objects.create(owner=self.owner, name='Regal')
        add_field(self.coll, 'name', 'Name', FieldType.TEXT)
        CollectionShare.objects.create(collection=self.coll, user=self.viewer,
                                       permission='view')

    def test_save_current_filters_drops_page_param(self):
        from .models import SavedView
        resp = self.client.post(reverse('saved_view_create', args=[self.coll.pk]),
                                {'name': 'Regal 3', 'querystring': 'q=krimi&page=4&sort=name'})
        self.assertEqual(resp.status_code, 302)
        view = SavedView.objects.get(collection=self.coll, name='Regal 3')
        self.assertIn('q=krimi', view.querystring)
        self.assertIn('sort=name', view.querystring)
        self.assertNotIn('page=', view.querystring)

    def test_saved_view_appears_as_link_and_reapplies(self):
        self.client.post(reverse('saved_view_create', args=[self.coll.pk]),
                         {'name': 'Krimis', 'querystring': 'q=krimi'})
        resp = self.client.get(reverse('collection_detail', args=[self.coll.pk]))
        self.assertContains(resp, 'Krimis')
        self.assertContains(resp, '?q=krimi')

    def test_same_name_updates_existing_view(self):
        from .models import SavedView
        create = reverse('saved_view_create', args=[self.coll.pk])
        self.client.post(create, {'name': 'A', 'querystring': 'q=x'})
        self.client.post(create, {'name': 'A', 'querystring': 'q=y'})
        views = SavedView.objects.filter(collection=self.coll, name='A')
        self.assertEqual(views.count(), 1)
        self.assertEqual(views.first().querystring, 'q=y')

    def test_viewer_cannot_manage_views(self):
        from .models import SavedView
        self.client.force_login(self.viewer)
        resp = self.client.post(reverse('saved_view_create', args=[self.coll.pk]),
                                {'name': 'X', 'querystring': ''})
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(SavedView.objects.exists())

    def test_delete_view(self):
        from .models import SavedView
        self.client.post(reverse('saved_view_create', args=[self.coll.pk]),
                         {'name': 'Weg', 'querystring': 'q=z'})
        view = SavedView.objects.get(name='Weg')
        self.client.post(reverse('saved_view_delete', args=[self.coll.pk, view.pk]))
        self.assertFalse(SavedView.objects.filter(pk=view.pk).exists())


@override_settings(MEDIA_ROOT=MEDIA)
class ItemGalleryTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('galowner', 'g@e.de', 'pw')
        self.viewer = User.objects.create_user('galviewer', 'gv@e.de', 'pw')
        self.client.force_login(self.owner)
        self.coll = Collection.objects.create(owner=self.owner, name='Galerie')
        add_field(self.coll, 'name', 'Name', FieldType.TEXT)
        CollectionShare.objects.create(collection=self.coll, user=self.viewer,
                                       permission='view')
        self.item = Item.objects.create(collection=self.coll, values={'name': 'Ding'})
        self.url = reverse('item_photo_add', args=[self.coll.pk, self.item.pk])

    def _photo(self, name='foto.png', color='blue'):
        return SimpleUploadedFile(name, make_png(color), 'image/png')

    def test_upload_multiple_photos(self):
        from .models import GALLERY_KEY
        resp = self.client.post(self.url, {'photos': [self._photo('a.png'),
                                                      self._photo('b.png', 'green')]})
        self.assertEqual(resp.status_code, 302)
        photos = self.item.assets.filter(field_key=GALLERY_KEY)
        self.assertEqual(photos.count(), 2)
        detail = self.client.get(reverse('item_detail', args=[self.coll.pk, self.item.pk]))
        self.assertContains(detail, 'bi-images')
        self.assertContains(detail, photos.first().file.url)

    def test_non_image_upload_is_rejected(self):
        from .models import GALLERY_KEY
        bad = SimpleUploadedFile('nicht-bild.png', b'kein bild', 'image/png')
        self.client.post(self.url, {'photos': [bad]})
        self.assertEqual(self.item.assets.filter(field_key=GALLERY_KEY).count(), 0)

    def test_viewer_cannot_upload(self):
        self.client.force_login(self.viewer)
        resp = self.client.post(self.url, {'photos': [self._photo()]})
        self.assertEqual(resp.status_code, 403)

    def test_delete_photo_removes_file_from_disk(self):
        import os
        from .models import GALLERY_KEY
        self.client.post(self.url, {'photos': [self._photo()]})
        asset = self.item.assets.get(field_key=GALLERY_KEY)
        path = asset.file.path
        self.assertTrue(os.path.exists(path))
        self.client.post(reverse('item_photo_delete',
                                 args=[self.coll.pk, self.item.pk, asset.pk]))
        self.assertFalse(os.path.exists(path))
        self.assertEqual(self.item.assets.count(), 0)

    def test_duplicate_copies_gallery_without_polluting_values(self):
        from .models import GALLERY_KEY
        self.client.post(self.url, {'photos': [self._photo()]})
        resp = self.client.post(reverse('item_duplicate',
                                        args=[self.coll.pk, self.item.pk]))
        self.assertEqual(resp.status_code, 302)
        copy = self.coll.items.exclude(pk=self.item.pk).get()
        self.assertEqual(copy.assets.filter(field_key=GALLERY_KEY).count(), 1)
        self.assertNotIn(GALLERY_KEY, copy.values)

    def test_field_keys_starting_with_underscore_are_rejected(self):
        from .forms import FieldDefinitionForm
        form = FieldDefinitionForm(
            {'label': 'Intern', 'key': '__gallery', 'field_type': 'text', 'order': 0},
            collection=self.coll)
        self.assertFalse(form.is_valid())
        self.assertIn('key', form.errors)


@override_settings(MEDIA_ROOT=MEDIA)
class BackupTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('bakowner', 'b@e.de', 'pw')
        self.other = User.objects.create_user('bakother', 'bo@e.de', 'pw')
        self.client.force_login(self.owner)
        self.coll = Collection.objects.create(owner=self.owner, name='Schatzkiste')
        add_field(self.coll, 'name', 'Name', FieldType.TEXT, 0)
        add_field(self.coll, 'bild', 'Bild', FieldType.IMAGE, 1)
        self.client.post(reverse('item_create', args=[self.coll.pk]), {
            'name': 'Goldmünze',
            'bild': SimpleUploadedFile('muenze.png', make_png(), 'image/png'),
        })
        self.item = self.coll.items.get()
        # One extra gallery photo — must land in the backup too.
        self.client.post(reverse('item_photo_add', args=[self.coll.pk, self.item.pk]),
                         {'photos': [SimpleUploadedFile('detail.png', make_png('green'),
                                                        'image/png')]})

    def test_backup_zip_contains_excel_json_and_media(self):
        import zipfile as ziplib
        resp = self.client.get(reverse('collection_backup', args=[self.coll.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/zip')
        self.assertIn('sicherung-schatzkiste', resp['Content-Disposition'])
        archive = ziplib.ZipFile(io.BytesIO(b''.join(resp.streaming_content)))
        names = archive.namelist()
        self.assertIn('daten.xlsx', names)
        self.assertIn('sammlung.json', names)
        media = [n for n in names if n.startswith('medien/')]
        self.assertEqual(len(media), 2)  # image field + gallery photo
        data = json.loads(archive.read('sammlung.json'))
        self.assertEqual(data['name'], 'Schatzkiste')
        self.assertEqual(data['items'][0]['values']['name'], 'Goldmünze')
        self.assertEqual(len(data['items'][0]['files']), 2)

    def test_backup_includes_trashed_items_in_json(self):
        self.item.soft_delete()
        import zipfile as ziplib
        resp = self.client.get(reverse('collection_backup', args=[self.coll.pk]))
        archive = ziplib.ZipFile(io.BytesIO(b''.join(resp.streaming_content)))
        data = json.loads(archive.read('sammlung.json'))
        self.assertIsNotNone(data['items'][0]['deleted_at'])
        self.assertTrue(any(n.startswith('medien/') for n in archive.namelist()))

    def test_backup_requires_access(self):
        self.client.force_login(self.other)
        resp = self.client.get(reverse('collection_backup', args=[self.coll.pk]))
        self.assertEqual(resp.status_code, 403)

    def test_backup_is_rate_limited(self):
        url = reverse('collection_backup', args=[self.coll.pk])
        for _ in range(10):
            self.assertEqual(self.client.get(url).status_code, 200)
        self.assertEqual(self.client.get(url).status_code, 429)

    def test_account_export_reuses_collection_json(self):
        resp = self.client.get(reverse('data_export'))
        data = json.loads(resp.content)
        collection = data['collections'][0]
        self.assertEqual(collection['name'], 'Schatzkiste')
        self.assertIn('saved_views', collection)


@override_settings(MEDIA_ROOT=MEDIA)
class RestoreTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('restowner', 'r@e.de', 'pw')
        self.friend = User.objects.create_user('restfriend', 'rf@e.de', 'pw')
        self.client.force_login(self.owner)
        self.coll = Collection.objects.create(owner=self.owner, name='Original',
                                              description='Meine Schätze',
                                              lookup_provider='books')
        self.f_name = add_field(self.coll, 'name', 'Name', FieldType.TEXT, 0)
        add_field(self.coll, 'genre', 'Genre', FieldType.CHOICE, 1,
                  config={'choices': ['Krimi', 'SF'], 'lookup_attribute': 'categories'})
        add_field(self.coll, 'bild', 'Bild', FieldType.IMAGE, 2)
        self.art = ItemType.objects.create(collection=self.coll, name='Buch')
        self.f_name.required_for_types.set([self.art])
        CollectionShare.objects.create(collection=self.coll, user=self.friend,
                                       permission='view')
        from .models import SavedView
        SavedView.objects.create(collection=self.coll, name='Krimis',
                                 querystring='f_genre=Krimi', created_by=self.owner)
        # Item with type, image, gallery photo and a loan.
        self.client.post(reverse('item_create', args=[self.coll.pk]), {
            '__item_type': self.art.pk, 'name': 'Der Alienist', 'genre': 'Krimi',
            'bild': SimpleUploadedFile('cover.png', make_png(), 'image/png'),
        })
        self.item = self.coll.items.get()
        self.client.post(reverse('item_photo_add', args=[self.coll.pk, self.item.pk]),
                         {'photos': [SimpleUploadedFile('back.png', make_png('green'),
                                                        'image/png')]})
        Loan.objects.create(item=self.item, borrower='Anna', note='vorsichtig!')
        # A second, trashed item.
        trashed = Item.objects.create(collection=self.coll, values={'name': 'Alt'})
        trashed.soft_delete()

    def _backup_bytes(self) -> bytes:
        resp = self.client.get(reverse('collection_backup', args=[self.coll.pk]))
        return b''.join(resp.streaming_content)

    def test_full_round_trip_restores_everything_but_shares(self):
        import os
        from .models import GALLERY_KEY, SavedView
        backup = self._backup_bytes()
        self.client.force_login(self.friend)
        resp = self.client.post(reverse('collection_restore'), {
            'file': SimpleUploadedFile('sicherung.zip', backup, 'application/zip')})
        self.assertEqual(resp.status_code, 302)
        restored = Collection.objects.get(owner=self.friend)
        self.assertEqual(restored.name, 'Original')
        self.assertEqual(restored.lookup_provider, 'books')
        # Structure: fields incl. config, item type, per-type required mapping.
        genre = restored.fields.get(key='genre')
        self.assertEqual(genre.config['choices'], ['Krimi', 'SF'])
        self.assertEqual(genre.config['lookup_attribute'], 'categories')
        art = restored.item_types.get()
        self.assertEqual(art.name, 'Buch')
        self.assertEqual(list(restored.fields.get(key='name').required_for_types.all()),
                         [art])
        # Saved view.
        self.assertEqual(SavedView.objects.get(collection=restored).querystring,
                         'f_genre=Krimi')
        # Items: one active with remapped file values, one in the trash.
        item = restored.items.get()
        self.assertEqual(item.values['name'], 'Der Alienist')
        self.assertEqual(item.item_type, art)
        asset_ref = item.values['bild']
        asset = item.assets.get(pk=asset_ref['asset_id'])
        self.assertNotEqual(str(asset.pk), str(self.item.assets.exclude(
            field_key=GALLERY_KEY).get().pk))
        self.assertTrue(os.path.exists(asset.file.path))
        self.assertEqual(item.assets.filter(field_key=GALLERY_KEY).count(), 1)
        self.assertNotIn(GALLERY_KEY, item.values)
        self.assertEqual(item.loans.get().borrower, 'Anna')
        self.assertEqual(Item.all_objects.filter(collection=restored,
                                                 deleted_at__isnull=False).count(), 1)
        # Shares are deliberately NOT restored.
        self.assertEqual(restored.shares.count(), 0)

    def test_garbage_zip_is_rejected(self):
        self.client.force_login(self.friend)
        resp = self.client.post(reverse('collection_restore'), {
            'file': SimpleUploadedFile('kaputt.zip', b'das ist kein zip', 'application/zip')})
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(Collection.objects.filter(owner=self.friend).exists())

    def test_zip_without_manifest_is_rejected(self):
        import zipfile as ziplib
        buf = io.BytesIO()
        with ziplib.ZipFile(buf, 'w') as zf:
            zf.writestr('irgendwas.txt', 'hallo')
        self.client.force_login(self.friend)
        self.client.post(reverse('collection_restore'), {
            'file': SimpleUploadedFile('x.zip', buf.getvalue(), 'application/zip')})
        self.assertFalse(Collection.objects.filter(owner=self.friend).exists())

    def test_member_count_guard(self):
        import zipfile as ziplib
        from . import restore as restore_mod
        buf = io.BytesIO()
        with ziplib.ZipFile(buf, 'w') as zf:
            for i in range(4):
                zf.writestr(f'file{i}.txt', 'x')
        self.client.force_login(self.friend)
        with mock.patch.object(restore_mod, 'MAX_MEMBERS', 3):
            self.client.post(reverse('collection_restore'), {
                'file': SimpleUploadedFile('bomb.zip', buf.getvalue(), 'application/zip')})
        self.assertFalse(Collection.objects.filter(owner=self.friend).exists())


class NotificationTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user('notifowner', 'no@e.de', 'pw')
        self.friend = User.objects.create_user('notiffriend', 'nf@e.de', 'pw')
        self.coll = Collection.objects.create(owner=self.owner, name='Geteiltes')

    def _share(self, permission='view'):
        self.client.force_login(self.owner)
        return self.client.post(reverse('collection_shares', args=[self.coll.pk]),
                                {'identifier': 'notiffriend', 'permission': permission})

    def test_share_creates_notification_for_recipient(self):
        from .models import Notification
        self._share()
        notification = Notification.objects.get(user=self.friend)
        self.assertEqual(notification.kind, Notification.KIND_SHARE)
        self.assertIn('Geteiltes', notification.message)
        self.assertIn('notifowner', notification.message)
        self.assertIsNone(notification.read_at)

    def test_bell_shows_unread_badge_and_message(self):
        self._share()
        self.client.force_login(self.friend)
        resp = self.client.get(reverse('dashboard'))
        self.assertContains(resp, 'bi-bell')
        self.assertContains(resp, 'Geteiltes')
        self.assertContains(resp, 'Alle als gelesen markieren')

    def test_open_marks_read_and_redirects_to_target(self):
        from .models import Notification
        self._share()
        notification = Notification.objects.get(user=self.friend)
        self.client.force_login(self.friend)
        resp = self.client.get(reverse('notification_open', args=[notification.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertIn(str(self.coll.pk), resp['Location'])
        notification.refresh_from_db()
        self.assertIsNotNone(notification.read_at)

    def test_cannot_open_foreign_notifications(self):
        from .models import Notification
        self._share()
        notification = Notification.objects.get(user=self.friend)
        self.client.force_login(self.owner)
        resp = self.client.get(reverse('notification_open', args=[notification.pk]))
        self.assertEqual(resp.status_code, 404)

    def test_reshare_resurfaces_read_notification(self):
        from .models import Notification
        self._share()
        Notification.objects.filter(user=self.friend).update(read_at='2026-01-01T00:00:00Z')
        self._share(permission='edit')
        notification = Notification.objects.get(user=self.friend)
        self.assertIsNone(notification.read_at)
        self.assertIn('Bearbeiten', notification.message)

    def test_share_revoke_removes_notification(self):
        from .models import Notification
        self._share()
        share = CollectionShare.objects.get()
        self.client.post(reverse('share_delete', args=[self.coll.pk, share.pk]))
        self.assertFalse(Notification.objects.filter(user=self.friend).exists())

    def test_mark_all_read(self):
        from .models import Notification
        self._share()
        self.client.force_login(self.friend)
        self.client.post(reverse('notifications_read_all'), {'next': '/'})
        self.assertFalse(Notification.objects.filter(user=self.friend,
                                                     read_at__isnull=True).exists())

    def test_pending_registration_notifies_staff(self):
        from .models import Notification
        staff = User.objects.create_user('notifstaff', 'ns@e.de', 'pw', is_staff=True)
        self.client.logout()
        self.client.post(reverse('register'), {
            'username': 'newbie', 'email': 'newbie@e.de',
            'password1': 'sicheres-passwort-77', 'password2': 'sicheres-passwort-77'})
        notification = Notification.objects.get(user=staff)
        self.assertEqual(notification.kind, Notification.KIND_REGISTRATION)
        self.assertIn('newbie', notification.message)

    def test_overdue_loans_appear_in_bell(self):
        from datetime import timedelta
        from django.utils import timezone
        item = Item.objects.create(collection=self.coll, values={'name': 'Buch'})
        Loan.objects.create(item=item, borrower='Bo',
                            lent_at=timezone.localdate() - timedelta(days=2),
                            due_at=timezone.localdate() - timedelta(days=1))
        self.client.force_login(self.owner)
        resp = self.client.get(reverse('collection_list'))
        self.assertContains(resp, 'überfällige Ausleihe')


class WorkflowTests(TestCase):
    """Item-type management, inline editing and the fast-capture button."""

    def setUp(self):
        self.owner = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.viewer = User.objects.create_user('viewer', 'v@e.de', 'pw')
        self.col = Collection.objects.create(owner=self.owner, name='Sammlung')
        CollectionShare.objects.create(collection=self.col, user=self.viewer,
                                        permission=CollectionShare.Permission.VIEW)
        add_field(self.col, 'name', 'Name', FieldType.TEXT, order=0)
        add_field(self.col, 'note', 'Notiz', FieldType.TEXT, order=1)
        add_field(self.col, 'count', 'Anzahl', FieldType.NUMBER, order=2)
        add_field(self.col, 'pic', 'Bild', FieldType.IMAGE, order=3)
        self.dvd = ItemType.objects.create(collection=self.col, name='DVD')
        self.client.force_login(self.owner)

    # --- item type edit/delete ---
    def test_type_edit_renames(self):
        resp = self.client.post(reverse('type_edit', args=[self.col.pk, self.dvd.pk]),
                                {'name': 'Blu-ray', 'order': 0})
        self.assertEqual(resp.status_code, 302)
        self.dvd.refresh_from_db()
        self.assertEqual(self.dvd.name, 'Blu-ray')

    def test_type_delete_unassigns_items(self):
        item = Item.objects.create(collection=self.col, item_type=self.dvd, values={'name': 'X'})
        resp = self.client.post(reverse('type_delete', args=[self.col.pk, self.dvd.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(ItemType.objects.filter(pk=self.dvd.pk).exists())
        item.refresh_from_db()
        self.assertIsNone(item.item_type_id)  # SET_NULL, item survives

    def test_type_delete_confirm_page_counts_items(self):
        Item.objects.create(collection=self.col, item_type=self.dvd, values={'name': 'X'})
        resp = self.client.get(reverse('type_delete', args=[self.col.pk, self.dvd.pk]))
        self.assertEqual(resp.status_code, 200)

    def test_viewer_cannot_delete_type(self):
        self.client.force_login(self.viewer)
        resp = self.client.post(reverse('type_delete', args=[self.col.pk, self.dvd.pk]))
        self.assertEqual(resp.status_code, 403)

    # --- save & next ---
    def test_save_and_new_returns_to_create_with_type(self):
        resp = self.client.post(reverse('item_create', args=[self.col.pk]),
                                {'__item_type': str(self.dvd.pk), 'name': 'A', 'save_and_new': '1'})
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse('item_create', args=[self.col.pk]), resp['Location'])
        self.assertIn(f'item_type={self.dvd.pk}', resp['Location'])
        self.assertEqual(Item.objects.filter(collection=self.col).count(), 1)

    def test_plain_save_returns_to_detail(self):
        resp = self.client.post(reverse('item_create', args=[self.col.pk]),
                                {'__item_type': str(self.dvd.pk), 'name': 'A'})
        self.assertRedirects(resp, reverse('collection_detail', args=[self.col.pk]))

    # --- inline editing ---
    def test_inline_update_changes_value(self):
        item = Item.objects.create(collection=self.col, values={'name': 'A'})
        resp = self.client.post(reverse('item_inline_update', args=[self.col.pk, item.pk]),
                                {'field_key': 'note', 'value': 'Hallo'})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()['ok'])
        item.refresh_from_db()
        self.assertEqual(item.values['note'], 'Hallo')

    def test_inline_update_clears_value_when_empty(self):
        item = Item.objects.create(collection=self.col, values={'name': 'A', 'note': 'X'})
        self.client.post(reverse('item_inline_update', args=[self.col.pk, item.pk]),
                         {'field_key': 'note', 'value': ''})
        item.refresh_from_db()
        self.assertNotIn('note', item.values)

    def test_inline_update_validates_number(self):
        item = Item.objects.create(collection=self.col, values={'name': 'A'})
        resp = self.client.post(reverse('item_inline_update', args=[self.col.pk, item.pk]),
                                {'field_key': 'count', 'value': 'keine-zahl'})
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.json()['ok'])

    def test_inline_update_rejects_file_field(self):
        item = Item.objects.create(collection=self.col, values={'name': 'A'})
        resp = self.client.post(reverse('item_inline_update', args=[self.col.pk, item.pk]),
                                {'field_key': 'pic', 'value': 'x'})
        self.assertEqual(resp.status_code, 400)

    def test_inline_update_unknown_field(self):
        item = Item.objects.create(collection=self.col, values={'name': 'A'})
        resp = self.client.post(reverse('item_inline_update', args=[self.col.pk, item.pk]),
                                {'field_key': 'nope', 'value': 'x'})
        self.assertEqual(resp.status_code, 400)

    def test_viewer_cannot_inline_update(self):
        item = Item.objects.create(collection=self.col, values={'name': 'A'})
        self.client.force_login(self.viewer)
        resp = self.client.post(reverse('item_inline_update', args=[self.col.pk, item.pk]),
                                {'field_key': 'note', 'value': 'X'})
        self.assertEqual(resp.status_code, 403)


DISCOGS_SEARCH = {'results': [{'id': 111, 'title': 'Radiohead - OK Computer',
                              'thumb': 'https://img/thumb.jpg'}]}
DISCOGS_STATS = {'num_for_sale': 3, 'lowest_price': {'value': 9.99, 'currency': 'EUR'}}


class OfferProviderTests(TestCase):
    def setUp(self):
        from . import runtime_settings
        runtime_settings.set_setting('live_offers_enabled', True)
        runtime_settings.set_setting('discogs_token', 'tok')
        self.user = User.objects.create_user('owner', 'o@e.de', 'pw')
        self.client.force_login(self.user)

    def _query(self, **kw):
        from .price_search import PriceQuery
        return PriceQuery(**kw)

    def test_discogs_offers_parsed(self):
        from . import offer_providers
        with mock.patch.object(lookup_providers, '_http_get_json',
                               side_effect=[DISCOGS_SEARCH, DISCOGS_STATS]):
            offers = offer_providers.fetch_offers(self._query(q='OK Computer', kind='music'))
        self.assertEqual(len(offers), 1)
        self.assertEqual(str(offers[0].price), '9.99')
        self.assertIn('Radiohead', offers[0].title)

    def test_no_offers_without_token(self):
        from . import offer_providers, runtime_settings
        runtime_settings.set_setting('discogs_token', '')
        with mock.patch.object(lookup_providers, '_http_get_json') as m:
            offers = offer_providers.fetch_offers(self._query(q='x', kind='music'))
        self.assertEqual(offers, [])
        m.assert_not_called()

    def test_provider_skips_non_music(self):
        from . import offer_providers
        with mock.patch.object(lookup_providers, '_http_get_json') as m:
            offers = offer_providers.fetch_offers(self._query(q='x', kind='books'))
        self.assertEqual(offers, [])
        m.assert_not_called()

    def test_price_page_shows_live_offers(self):
        with mock.patch.object(lookup_providers, '_http_get_json',
                               side_effect=[DISCOGS_SEARCH, DISCOGS_STATS]):
            resp = self.client.get(reverse('price_search'), {'q': 'OK Computer', 'kind': 'music'})
        self.assertContains(resp, 'Live-Angebote')
        self.assertContains(resp, '9.99')

    def test_price_page_hides_offers_when_disabled(self):
        from . import runtime_settings
        runtime_settings.set_setting('live_offers_enabled', False)
        resp = self.client.get(reverse('price_search'), {'q': 'OK Computer', 'kind': 'music'})
        self.assertNotContains(resp, 'Live-Angebote')
