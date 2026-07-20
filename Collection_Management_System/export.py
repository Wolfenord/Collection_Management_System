"""Dynamic exports of a collection.

* ``build_workbook`` — Excel (.xlsx): columns from the FieldDefinitions, rows
  from whatever items are passed in (the view feeds the *filtered* queryset).
* ``collection_json`` — complete machine-readable dump of one collection
  (structure + contents incl. trash); shared by the GDPR account export and
  the ZIP backup.
* ``build_backup_zip`` — full backup: Excel + JSON + every uploaded file.
"""

from __future__ import annotations

import json
import os
import re
import zipfile
from io import BytesIO
from tempfile import SpooledTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.utils.translation import gettext as _

from .codes import item_short_code
from .models import FieldType

_INVALID_SHEET = re.compile(r'[\\/?*\[\]:]')
NUMERIC_TYPES = {FieldType.NUMBER, FieldType.YEAR, FieldType.DECIMAL, FieldType.PRICE}


def _cell_value(field, raw, build_uri):
    """Convert a stored value into something Excel should display."""
    if raw in (None, '', [], {}):
        return ''
    t = field.field_type
    if t in (FieldType.IMAGE, FieldType.FILE) and isinstance(raw, dict):
        url = raw.get('url', '')
        return build_uri(url) if url else raw.get('name', '')
    if t == FieldType.BOOLEAN:
        return _('Ja') if raw else _('Nein')
    if t == FieldType.MULTICHOICE and isinstance(raw, list):
        return ', '.join(str(v) for v in raw)
    if t in NUMERIC_TYPES:
        try:
            return float(raw) if t in (FieldType.DECIMAL, FieldType.PRICE) else int(raw)
        except (TypeError, ValueError):
            return str(raw)
    return str(raw)


def build_workbook(collection, items, fields, build_uri) -> bytes:
    """Return xlsx bytes: one column per field, one row per item."""
    wb = Workbook()
    ws = wb.active
    ws.title = (_INVALID_SHEET.sub('', collection.name) or 'Export')[:31]

    headers = [_('ID'), _('Art')] + [f.label for f in fields]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'

    for item in items:
        values = item.values or {}
        row = [item_short_code(item), str(item.item_type) if item.item_type else '']
        row += [_cell_value(f, values.get(f.key), build_uri) for f in fields]
        ws.append(row)

    # Rough auto-width based on header/content length.
    for idx, header in enumerate(headers, start=1):
        width = max(len(str(header)), 12)
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
            width = max(width, len(str(row[0])) if row[0] is not None else 0)
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def collection_json(collection) -> dict:
    """One collection with its complete structure and contents (incl. trash)."""
    from .models import Item

    return {
        'id': str(collection.pk),
        'name': collection.name,
        'description': collection.description,
        'media_kind': collection.lookup_provider,
        'created_at': collection.created_at.isoformat(),
        'fields': [
            {'key': fd.key, 'label': str(fd.label), 'type': fd.field_type,
             'required': fd.required, 'config': fd.config,
             'required_for': [t.name for t in fd.required_for_types.all()]}
            for fd in collection.fields.all().prefetch_related('required_for_types')
        ],
        'item_types': [{'name': it.name, 'description': it.description}
                       for it in collection.item_types.all()],
        'shared_with': [
            {'user': share.user.get_username(), 'permission': share.permission}
            for share in collection.shares.select_related('user')
        ],
        'saved_views': [{'name': view.name, 'querystring': view.querystring}
                        for view in collection.saved_views.all()],
        'items': [
            {
                'id': str(item.pk),
                'type': item.item_type.name if item.item_type else None,
                'created_at': item.created_at.isoformat(),
                'deleted_at': item.deleted_at.isoformat() if item.deleted_at else None,
                'values': item.values,
                'files': [{'field': a.field_key, 'name': a.original_name, 'path': a.file.name}
                          for a in item.assets.all()],
                'loans': [
                    {'borrower': loan.borrower, 'lent_at': str(loan.lent_at),
                     'due_at': str(loan.due_at) if loan.due_at else None,
                     'returned_at': str(loan.returned_at) if loan.returned_at else None,
                     'note': loan.note}
                    for loan in item.loans.all()
                ],
            }
            for item in Item.all_objects.filter(collection=collection)
            .select_related('item_type').prefetch_related('assets', 'loans')
        ],
    }


def build_backup_zip(collection, build_uri):
    """Full backup of one collection as a ZIP file object (seeked to 0).

    Contents: ``daten.xlsx`` (current items, like the export), ``sammlung.json``
    (complete machine-readable dump incl. trash) and ``medien/<item>/<file>``
    for every uploaded file — the pieces needed to rebuild or migrate the
    collection. Spooled to disk beyond 32 MB so huge media sets don't live in
    memory.
    """
    from .models import Item

    fields = list(collection.fields.all())
    items = list(collection.items.select_related('item_type'))
    buffer = SpooledTemporaryFile(max_size=32 * 1024 * 1024)
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
        archive.writestr('daten.xlsx', build_workbook(collection, items, fields, build_uri))
        archive.writestr('sammlung.json',
                         json.dumps(collection_json(collection), ensure_ascii=False, indent=2))
        seen: set[str] = set()
        for item in (Item.all_objects.filter(collection=collection)
                     .prefetch_related('assets')):
            short = item.id.hex[:12]
            for asset in item.assets.all():
                base = os.path.basename(asset.file.name) or 'datei'
                name = f'medien/{short}/{base}'
                if name in seen:  # same stored filename twice on one item
                    name = f'medien/{short}/{asset.id.hex[:8]}-{base}'
                seen.add(name)
                try:
                    with asset.file.open('rb') as fh:
                        archive.writestr(name, fh.read())
                except (OSError, ValueError):
                    continue  # file missing on disk: keep the rest of the backup
    buffer.seek(0)
    return buffer
